import openpyxl

# Load the Excel file
wb = openpyxl.load_workbook('/app/data/templumis_university.xlsx', data_only=True)

print("Available sheets:", wb.sheetnames)
print("\n" + "="*60)

# Check if Rankings Dashboard exists
if 'Rankings Dashboard' in wb.sheetnames:
    ws = wb['Rankings Dashboard']
    print("Found 'Rankings Dashboard' sheet!")
elif 'Dashboard' in wb.sheetnames:
    ws = wb['Dashboard']
    print("Found 'Dashboard' sheet - checking content...")
else:
    print("No Rankings Dashboard found")
    exit()

print("\nFirst 20 rows of data:")
print("="*60)

for idx, row in enumerate(ws.iter_rows(min_row=1, max_row=20, values_only=True), 1):
    print(f"Row {idx}: {row}")

print("\n" + "="*60)
print("Sheet dimensions:", ws.dimensions)
