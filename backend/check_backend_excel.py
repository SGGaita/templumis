import openpyxl

# Check both possible locations
files_to_check = [
    '/app/data/templumis_university.xlsx',
    'data/templumis_university.xlsx'
]

for file_path in files_to_check:
    try:
        print(f"\n{'='*60}")
        print(f"Checking: {file_path}")
        print('='*60)
        
        wb = openpyxl.load_workbook(file_path, data_only=True)
        print(f"✓ File found!")
        print(f"Sheets: {wb.sheetnames}")
        
        # Look for Rankings Dashboard
        if 'Rankings Dashboard' in wb.sheetnames:
            print("\n⭐ FOUND 'Rankings Dashboard' sheet!")
            ws = wb['Rankings Dashboard']
            print(f"Dimensions: {ws.dimensions}")
            print("\nFirst 30 rows:")
            for idx, row in enumerate(ws.iter_rows(min_row=1, max_row=30, values_only=True), 1):
                if any(cell is not None for cell in row):
                    print(f"Row {idx}: {row}")
        else:
            print("❌ No 'Rankings Dashboard' sheet found")
            
    except FileNotFoundError:
        print(f"✗ File not found at {file_path}")
    except Exception as e:
        print(f"✗ Error: {e}")
