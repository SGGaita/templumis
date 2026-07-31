from fastapi import APIRouter, Depends, HTTPException, status
from sqlalchemy.orm import Session
from typing import List
from datetime import datetime, date
import openpyxl

from app.database import get_db
from app.auth import get_current_user
from app.models import (
    User, Student, StudentJourneyMilestone, StudentAdvisor,
    StudentRiskScore, PostgradSupport
)
from app.schemas import (
    JourneyMilestoneCreate, JourneyMilestoneUpdate, JourneyMilestoneOut,
    StudentAdvisorOut, RiskScoreOut, PostgradSupportOut
)

router = APIRouter(prefix="/api/student-journey", tags=["Student Journey"])

from app.excel_paths import resolve_excel_path
from app.pg_excel import (
    load_library_resources,
    load_phd_research,
    load_program_advisors,
    load_pg_research,
    load_pg_support,
    pg_alerts_from_research,
    pg_alerts_from_support,
)

def _journey_excel_path():
    return str(resolve_excel_path())


def _is_postgraduate_type(student_type: str, programme_level: str) -> bool:
    blob = f"{student_type} {programme_level}".lower()
    return any(k in blob for k in ("post", "master", "msc", "mba", "ma ", "mphil", "phd", "doctor"))


@router.get("/my-journey", response_model=dict)
async def get_my_journey(
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Get current student's journey timeline with milestones, advisor, and risk info"""
    
    if current_user.account_category != "student":
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Only students can access their journey"
        )
    
    # Get student record
    student = db.query(Student).filter(
        Student.institution_id == current_user.institution_id,
        Student.email == current_user.email
    ).first()
    
    if not student:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Student record not found"
        )
    
    # Get milestones
    milestones = db.query(StudentJourneyMilestone).filter(
        StudentJourneyMilestone.student_id == student.id
    ).order_by(StudentJourneyMilestone.milestone_date).all()
    
    # Get active advisor
    advisor_assignment = db.query(StudentAdvisor).filter(
        StudentAdvisor.student_id == student.id,
        StudentAdvisor.is_active == True
    ).first()
    
    advisor_info = None
    if advisor_assignment:
        advisor = db.query(User).filter(User.id == advisor_assignment.advisor_id).first()
        if advisor:
            advisor_info = {
                "id": advisor.id,
                "name": advisor.full_name,
                "email": advisor.email,
                "advisor_type": advisor_assignment.advisor_type,
                "assignment_date": advisor_assignment.assignment_date.isoformat() if advisor_assignment.assignment_date else None
            }
    
    # Get latest risk score
    latest_risk = db.query(StudentRiskScore).filter(
        StudentRiskScore.student_id == student.id
    ).order_by(StudentRiskScore.calculated_at.desc()).first()
    
    risk_info = None
    if latest_risk:
        risk_info = {
            "risk_score": latest_risk.risk_score,
            "risk_level": latest_risk.risk_level,
            "risk_factors": latest_risk.risk_factors,
            "calculated_at": latest_risk.calculated_at.isoformat()
        }
    
    # Get postgrad support if applicable
    postgrad_info = None
    if student.program_level in ["masters", "phd", "postgraduate"]:
        postgrad = db.query(PostgradSupport).filter(
            PostgradSupport.student_id == student.id
        ).first()
        if postgrad:
            postgrad_info = {
                "research_area": postgrad.research_area,
                "thesis_status": postgrad.thesis_status,
                "publications_count": len(postgrad.publications) if postgrad.publications else 0,
                "conferences_count": len(postgrad.conference_attendance) if postgrad.conference_attendance else 0
            }
    
    return {
        "student": {
            "id": student.id,
            "name": student.full_name,
            "student_number": student.student_number,
            "program_level": student.program_level,
            "enrollment_date": student.enrollment_date.isoformat() if student.enrollment_date else None,
            "expected_graduation": student.expected_graduation.isoformat() if student.expected_graduation else None,
            "current_milestone": student.current_milestone,
            "gpa": float(student.gpa) if student.gpa else None,
            "credits_completed": student.credits_completed,
            "status": student.status
        },
        "milestones": [
            {
                "id": m.id,
                "milestone_type": m.milestone_type,
                "milestone_date": m.milestone_date.isoformat(),
                "status": m.status,
                "notes": m.notes
            } for m in milestones
        ],
        "advisor": advisor_info,
        "risk_info": risk_info,
        "postgrad_info": postgrad_info
    }


@router.get("/milestones", response_model=List[JourneyMilestoneOut])
async def get_my_milestones(
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Get all milestones for current student"""
    
    if current_user.account_category != "student":
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Only students can access milestones"
        )
    
    student = db.query(Student).filter(
        Student.institution_id == current_user.institution_id,
        Student.email == current_user.email
    ).first()
    
    if not student:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Student record not found"
        )
    
    milestones = db.query(StudentJourneyMilestone).filter(
        StudentJourneyMilestone.student_id == student.id
    ).order_by(StudentJourneyMilestone.milestone_date).all()
    
    return milestones


@router.get("/progress/{student_id}", response_model=dict)
async def get_student_progress(
    student_id: int,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Admin view of student progress (staff only)"""
    
    if current_user.account_category != "staff":
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Only staff can view student progress"
        )
    
    student = db.query(Student).filter(
        Student.id == student_id,
        Student.institution_id == current_user.institution_id
    ).first()
    
    if not student:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Student not found"
        )
    
    # Get all milestones
    milestones = db.query(StudentJourneyMilestone).filter(
        StudentJourneyMilestone.student_id == student.id
    ).order_by(StudentJourneyMilestone.milestone_date).all()
    
    # Get advisors
    advisors = db.query(StudentAdvisor).filter(
        StudentAdvisor.student_id == student.id
    ).all()
    
    # Get risk scores history
    risk_scores = db.query(StudentRiskScore).filter(
        StudentRiskScore.student_id == student.id
    ).order_by(StudentRiskScore.calculated_at.desc()).limit(10).all()
    
    return {
        "student": {
            "id": student.id,
            "name": student.full_name,
            "student_number": student.student_number,
            "email": student.email,
            "program_level": student.program_level,
            "enrollment_date": student.enrollment_date.isoformat() if student.enrollment_date else None,
            "expected_graduation": student.expected_graduation.isoformat() if student.expected_graduation else None,
            "current_milestone": student.current_milestone,
            "gpa": float(student.gpa) if student.gpa else None,
            "credits_completed": student.credits_completed,
            "status": student.status,
            "risk_level": student.risk_level
        },
        "milestones": [
            {
                "id": m.id,
                "milestone_type": m.milestone_type,
                "milestone_date": m.milestone_date.isoformat(),
                "status": m.status,
                "notes": m.notes
            } for m in milestones
        ],
        "advisors": [
            {
                "id": a.id,
                "advisor_id": a.advisor_id,
                "advisor_type": a.advisor_type,
                "is_active": a.is_active,
                "assignment_date": a.assignment_date.isoformat() if a.assignment_date else None
            } for a in advisors
        ],
        "risk_history": [
            {
                "risk_score": r.risk_score,
                "risk_level": r.risk_level,
                "risk_factors": r.risk_factors,
                "calculated_at": r.calculated_at.isoformat()
            } for r in risk_scores
        ]
    }


@router.post("/milestones/{milestone_id}/complete")
async def complete_milestone(
    milestone_id: int,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Mark a milestone as complete (staff only)"""
    
    if current_user.account_category != "staff":
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Only staff can update milestones"
        )
    
    milestone = db.query(StudentJourneyMilestone).filter(
        StudentJourneyMilestone.id == milestone_id
    ).first()
    
    if not milestone:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Milestone not found"
        )
    
    # Verify institution
    student = db.query(Student).filter(Student.id == milestone.student_id).first()
    if student.institution_id != current_user.institution_id:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Cannot update milestone from different institution"
        )
    
    milestone.status = "completed"
    milestone.updated_at = datetime.utcnow()
    db.commit()
    
    return {"message": "Milestone marked as complete", "milestone_id": milestone_id}


@router.get("/my-journey-excel")
async def get_my_journey_from_excel(
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Get current student's journey from Excel Journey Tracker"""
    
    if current_user.account_category != "student":
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Only students can access their journey"
        )
    
    try:
        wb = openpyxl.load_workbook(_journey_excel_path(), data_only=True)
        ws = wb['Journey Tracker']
        
        # Get student registration number from user
        student_reg = current_user.student_registration_number
        if not student_reg:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Student registration number not found"
            )
        
        # Search for student in Journey Tracker
        student_journey = None
        for row in ws.iter_rows(min_row=3, values_only=True):
            if not row[0]:
                continue
            
            if str(row[0]) == str(student_reg):
                # Found the student
                academic_standing = str(row[12]) if row[12] else "Unknown"
                fees_balance = row[16] if row[16] else 0
                journey_stage = str(row[24]) if row[24] else "Unknown"
                gpa = row[11] if row[11] else 0
                
                # Calculate risk
                risk_level = "low"
                risk_score = 0
                risk_factors = []
                
                if academic_standing in ["Probation", "Suspended"]:
                    risk_score += 40
                    risk_factors.append("academic_standing")
                    risk_level = "critical"
                
                if fees_balance and fees_balance > 50000:
                    risk_score += 30
                    risk_factors.append("fee_arrears")
                    if risk_level != "critical":
                        risk_level = "high"
                elif fees_balance and fees_balance > 0:
                    risk_score += 15
                    risk_factors.append("fee_balance")
                    if risk_level == "low":
                        risk_level = "medium"
                
                if gpa and gpa < 2.0:
                    risk_score += 30
                    risk_factors.append("low_gpa")
                    risk_level = "critical"
                elif gpa and gpa < 2.5:
                    risk_score += 15
                    risk_factors.append("gpa_warning")
                    if risk_level == "low":
                        risk_level = "medium"
                
                # Generate alerts based on risk factors
                alerts = []
                if "academic_standing" in risk_factors:
                    alerts.append({
                        "type": "critical",
                        "title": "Academic Standing Alert",
                        "message": f"You are currently on {academic_standing}. Please contact your academic advisor.",
                        "action": "Contact Advisor"
                    })
                if "fee_arrears" in risk_factors:
                    alerts.append({
                        "type": "urgent",
                        "title": "Fee Payment Required",
                        "message": f"Outstanding balance: KES {fees_balance:,.0f}. Payment required to avoid registration block.",
                        "action": "Pay Fees"
                    })
                if "low_gpa" in risk_factors or "gpa_warning" in risk_factors:
                    alerts.append({
                        "type": "warning",
                        "title": "Academic Performance",
                        "message": f"Current GPA: {gpa:.2f}. Academic support services are available.",
                        "action": "Get Support"
                    })
                
                # Fetch additional data from other sheets
                # Helper function to convert sheet to dict list
                def sheet_to_dict_list(sheet):
                    data = []
                    headers = [cell.value for cell in sheet[1]]
                    for row in sheet.iter_rows(min_row=2, values_only=True):
                        if any(cell is not None for cell in row):
                            row_dict = {}
                            for idx, header in enumerate(headers):
                                if header and idx < len(row):
                                    row_dict[header] = row[idx]
                            data.append(row_dict)
                    return data
                
                # Get enrollments for this student
                enrollments = []
                courses_data = []
                grades_data = []
                attendance_data = []
                
                if "Enrolments" in wb.sheetnames:
                    enrolments_sheet = wb["Enrolments"]
                    all_enrolments = sheet_to_dict_list(enrolments_sheet)
                    enrollments = [e for e in all_enrolments if str(e.get("student_id")) == str(student_reg)]
                
                # Get courses
                if "Courses" in wb.sheetnames:
                    courses_sheet = wb["Courses"]
                    all_courses = sheet_to_dict_list(courses_sheet)
                    # Get courses for this student's enrollments
                    enrolled_course_ids = [e.get("course_id") for e in enrollments]
                    courses_data = [c for c in all_courses if c.get("course_id") in enrolled_course_ids]
                
                # Get grades
                if "Grades" in wb.sheetnames:
                    grades_sheet = wb["Grades"]
                    all_grades = sheet_to_dict_list(grades_sheet)
                    grades_data = [g for g in all_grades if str(g.get("student_id")) == str(student_reg)]
                
                # Get attendance
                if "Attendance" in wb.sheetnames:
                    attendance_sheet = wb["Attendance"]
                    all_attendance = sheet_to_dict_list(attendance_sheet)
                    attendance_data = [a for a in all_attendance if str(a.get("student_id")) == str(student_reg)]
                
                # Calculate attendance percentage
                total_sessions = len(attendance_data)
                attended_sessions = len([a for a in attendance_data if a.get("status") == "Present"])
                attendance_percentage = (attended_sessions / total_sessions * 100) if total_sessions > 0 else 0
                
                # Organize courses by semester/year
                courses_by_semester = {}
                for enrollment in enrollments:
                    semester = enrollment.get("semester", "Unknown")
                    course_id = enrollment.get("course_id")
                    
                    # Find course details
                    course_info = next((c for c in courses_data if c.get("course_id") == course_id), None)
                    
                    # Find grade for this course
                    grade_info = next((g for g in grades_data if g.get("course_id") == course_id), None)
                    
                    if semester not in courses_by_semester:
                        courses_by_semester[semester] = []
                    
                    courses_by_semester[semester].append({
                        "course_id": course_id,
                        "course_name": course_info.get("course_name") if course_info else "Unknown",
                        "credits": course_info.get("credits") if course_info else 0,
                        "grade": grade_info.get("grade") if grade_info else "N/A",
                        "score": grade_info.get("score") if grade_info else None,
                        "semester": semester
                    })
                
                student_type = str(row[2]) if row[2] else "Undergraduate"
                programme_level = str(row[3]) if row[3] else "BSc"
                is_postgrad = _is_postgraduate_type(student_type, programme_level)

                program_advisors = load_program_advisors(wb, student_reg)
                phd_research = load_phd_research(wb, student_reg) if is_postgrad else None
                pg_research = load_pg_research(wb, student_reg) if is_postgrad else None
                if is_postgrad and phd_research and not pg_research:
                    pg_research = {
                        "supervisor": phd_research.get("principal_supervisor"),
                        "co_supervisor": phd_research.get("co_supervisor"),
                        "dissertation_title": phd_research.get("dissertation_title"),
                    }
                elif is_postgrad and phd_research and pg_research:
                    if not pg_research.get("supervisor"):
                        pg_research["supervisor"] = phd_research.get("principal_supervisor")
                    if not pg_research.get("co_supervisor"):
                        pg_research["co_supervisor"] = phd_research.get("co_supervisor")
                pg_support = load_pg_support(wb, student_reg) if is_postgrad else []
                library_resources = load_library_resources(
                    wb, programme_level=programme_level, student_type=student_type
                )

                if is_postgrad:
                    if pg_research:
                        for alert in pg_alerts_from_research(pg_research):
                            alerts.append(alert)
                    for alert in pg_alerts_from_support(pg_support):
                        alerts.append(alert)

                student_journey = {
                    "student_id": str(row[0]),
                    "full_name": str(row[1]),
                    "student_type": student_type,
                    "programme_level": programme_level,
                    "cohort_level": "postgraduate" if is_postgrad else "undergraduate",
                    "department": str(row[4]) if row[4] else "Unknown",
                    "enrolment_date": str(row[5]) if row[5] else None,
                    "current_year_sem": str(row[6]) if row[6] else "Unknown",
                    "academic_progression": {
                        "sem_1_status": str(row[7]) if row[7] else "Pending",
                        "sem_2_status": str(row[8]) if row[8] else "Pending",
                        "year_2_progression": str(row[9]) if row[9] else "Pending",
                        "year_3_progression": str(row[10]) if row[10] else "Pending"
                    },
                    "academic_standing": {
                        "gpa": float(gpa) if gpa else 0.0,
                        "standing": academic_standing,
                        "last_review_date": str(row[13]) if row[13] else None
                    },
                    "financial_clearance": {
                        "fees_sem_1": str(row[14]) if row[14] else "Unknown",
                        "fees_sem_2": str(row[15]) if row[15] else "Unknown",
                        "fees_balance": float(fees_balance) if fees_balance else 0.0
                    },
                    "research_thesis": {
                        "proposal_status": str(row[17]) if row[17] else "N/A",
                        "data_coursework_phase": str(row[18]) if row[18] else "N/A",
                        "thesis_submission": str(row[19]) if row[19] else "N/A",
                        "thesis_defence": str(row[20]) if row[20] else "N/A"
                    },
                    "graduation": {
                        "clearance_status": str(row[21]) if row[21] else "Not Yet",
                        "clearance_date": str(row[22]) if row[22] else None,
                        "graduation_date": str(row[23]) if row[23] else None
                    },
                    "journey_stage": journey_stage,
                    "risk_assessment": {
                        "risk_score": risk_score,
                        "risk_level": risk_level,
                        "risk_factors": risk_factors
                    },
                    "alerts": alerts,
                    "courses_by_semester": courses_by_semester,
                    "total_courses": len(enrollments),
                    "total_credits": sum([c.get("credits", 0) for c in courses_data]),
                    "attendance_percentage": round(attendance_percentage, 1),
                    "total_attendance_sessions": total_sessions,
                    "attended_sessions": attended_sessions,
                    "program_advisors": program_advisors,
                    "phd_research": phd_research,
                    "pg_research": pg_research,
                    "pg_academic_support": pg_support,
                    "library_resources": library_resources,
                }
                break
        
        wb.close()
        
        if not student_journey:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Student journey data not found in Journey Tracker"
            )
        
        return student_journey
        
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error reading journey data: {str(e)}"
        )


@router.get("/library-resources")
async def get_library_resources(
    current_user: User = Depends(get_current_user),
):
    """FAIR-compliant library catalogue for the logged-in student."""
    if current_user.account_category != "student":
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Students only")

    student_reg = current_user.student_registration_number
    if not student_reg:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Student registration not found")

    try:
        wb = openpyxl.load_workbook(_journey_excel_path(), data_only=True)
        programme_level = "All"
        student_type = "Undergraduate"
        if "Journey Tracker" in wb.sheetnames:
            ws = wb["Journey Tracker"]
            for row in ws.iter_rows(min_row=3, values_only=True):
                if str(row[0]) == str(student_reg):
                    student_type = str(row[2] or "Undergraduate")
                    programme_level = str(row[3] or "BSc")
                    break
        resources = load_library_resources(
            wb, programme_level=programme_level, student_type=student_type
        )
        wb.close()
        return {"resources": resources, "total": len(resources)}
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error loading library resources: {str(e)}",
        ) from e


@router.get("/all-journeys")
async def get_all_student_journeys(
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Get all student journeys from Excel Journey Tracker (staff only)"""
    
    if current_user.account_category != "staff":
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Only staff can view all student journeys"
        )
    
    try:
        wb = openpyxl.load_workbook(_journey_excel_path(), data_only=True)
        ws = wb['Journey Tracker']
        
        journeys = []
        
        # Skip header rows (rows 1-2)
        for row in ws.iter_rows(min_row=3, values_only=True):
            if not row[0]:  # Skip empty rows
                continue
            
            # Calculate risk score based on academic standing and fees
            academic_standing = str(row[12]) if row[12] else "Unknown"
            fees_balance = row[16] if row[16] else 0
            journey_stage = str(row[24]) if row[24] else "Unknown"
            
            risk_level = "low"
            risk_score = 0
            risk_factors = []
            
            # Risk calculation
            if academic_standing in ["Probation", "Suspended"]:
                risk_score += 40
                risk_factors.append("academic_standing")
                risk_level = "critical"
            elif academic_standing == "Cleared":
                risk_score = 0
                risk_level = "low"
            
            if fees_balance and fees_balance > 50000:
                risk_score += 30
                risk_factors.append("fee_arrears")
                if risk_level != "critical":
                    risk_level = "high"
            elif fees_balance and fees_balance > 0:
                risk_score += 15
                risk_factors.append("fee_balance")
                if risk_level == "low":
                    risk_level = "medium"
            
            gpa = row[11] if row[11] else 0
            if gpa and gpa < 2.0:
                risk_score += 30
                risk_factors.append("low_gpa")
                risk_level = "critical"
            elif gpa and gpa < 2.5:
                risk_score += 15
                risk_factors.append("gpa_warning")
                if risk_level == "low":
                    risk_level = "medium"
            
            # Determine interventions needed
            interventions_needed = []
            if "academic_standing" in risk_factors:
                interventions_needed.append("Academic Counseling")
            if "fee_arrears" in risk_factors or "fee_balance" in risk_factors:
                interventions_needed.append("Financial Aid Review")
            if "low_gpa" in risk_factors or "gpa_warning" in risk_factors:
                interventions_needed.append("Academic Support")
            
            journey_data = {
                "student_id": str(row[0]),
                "full_name": str(row[1]),
                "student_type": str(row[2]) if row[2] else "Undergraduate",
                "programme_level": str(row[3]) if row[3] else "BSc",
                "department": str(row[4]) if row[4] else "Unknown",
                "enrolment_date": str(row[5]) if row[5] else None,
                "current_year_sem": str(row[6]) if row[6] else "Unknown",
                "academic_progression": {
                    "sem_1_status": str(row[7]) if row[7] else "Pending",
                    "sem_2_status": str(row[8]) if row[8] else "Pending",
                    "year_2_progression": str(row[9]) if row[9] else "Pending",
                    "year_3_progression": str(row[10]) if row[10] else "Pending"
                },
                "academic_standing": {
                    "gpa": float(gpa) if gpa else 0.0,
                    "standing": academic_standing,
                    "last_review_date": str(row[13]) if row[13] else None
                },
                "financial_clearance": {
                    "fees_sem_1": str(row[14]) if row[14] else "Unknown",
                    "fees_sem_2": str(row[15]) if row[15] else "Unknown",
                    "fees_balance": float(fees_balance) if fees_balance else 0.0
                },
                "research_thesis": {
                    "proposal_status": str(row[17]) if row[17] else "N/A",
                    "data_coursework_phase": str(row[18]) if row[18] else "N/A",
                    "thesis_submission": str(row[19]) if row[19] else "N/A",
                    "thesis_defence": str(row[20]) if row[20] else "N/A"
                },
                "graduation": {
                    "clearance_status": str(row[21]) if row[21] else "Not Yet",
                    "clearance_date": str(row[22]) if row[22] else None,
                    "graduation_date": str(row[23]) if row[23] else None
                },
                "journey_stage": journey_stage,
                "risk_assessment": {
                    "risk_score": risk_score,
                    "risk_level": risk_level,
                    "risk_factors": risk_factors,
                    "interventions_needed": interventions_needed
                }
            }
            
            journeys.append(journey_data)
        
        wb.close()
        
        # Calculate summary statistics
        total_students = len(journeys)
        risk_summary = {
            "critical": len([j for j in journeys if j["risk_assessment"]["risk_level"] == "critical"]),
            "high": len([j for j in journeys if j["risk_assessment"]["risk_level"] == "high"]),
            "medium": len([j for j in journeys if j["risk_assessment"]["risk_level"] == "medium"]),
            "low": len([j for j in journeys if j["risk_assessment"]["risk_level"] == "low"])
        }
        
        journey_stages = {}
        for j in journeys:
            stage = j["journey_stage"]
            journey_stages[stage] = journey_stages.get(stage, 0) + 1
        
        return {
            "total_students": total_students,
            "risk_summary": risk_summary,
            "journey_stages": journey_stages,
            "journeys": journeys
        }
        
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error reading journey data: {str(e)}"
        )
