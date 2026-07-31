from fastapi import APIRouter, WebSocket, WebSocketDisconnect
from watchdog.observers import Observer
from watchdog.events import FileSystemEventHandler
import asyncio
import json
import openpyxl
from typing import Set
import os

router = APIRouter()

# Store active WebSocket connections
active_connections: Set[WebSocket] = set()

# File path to watch
EXCEL_FILE_PATH = "/app/data/templumis_university.xlsx"


class ExcelFileHandler(FileSystemEventHandler):
    """Handler for Excel file changes"""
    
    def __init__(self, callback):
        self.callback = callback
        self.last_modified = 0
        
    def on_modified(self, event):
        if event.src_path.endswith("templumis_university.xlsx"):
            # Debounce: only trigger if more than 1 second has passed
            current_time = asyncio.get_event_loop().time()
            if current_time - self.last_modified > 1:
                self.last_modified = current_time
                print(f"📊 Excel file changed: {event.src_path}")
                asyncio.create_task(self.callback())


async def broadcast_update():
    """Broadcast update to all connected clients"""
    if not active_connections:
        return
    
    try:
        # Read updated data from Excel
        wb = openpyxl.load_workbook(EXCEL_FILE_PATH, data_only=True)
        ws = wb['Rankings Dashboard']
        
        # Parse institutional summary
        institutional_data = {
            "total_students": int(ws['B5'].value) if ws['B5'].value else 37,
            "ug_students": int(ws['B6'].value) if ws['B6'].value else 25,
            "pg_students": int(ws['B7'].value) if ws['B7'].value else 12,
            "faculty": int(ws['B8'].value) if ws['B8'].value else 15,
            "avg_gpa": str(ws['D5'].value) if ws['D5'].value else "3.32 / 4.0",
            "international_students": str(ws['F5'].value) if ws['F5'].value else "35.1%",
            "female_ratio": str(ws['H5'].value) if ws['H5'].value else "48.6%",
            "research_students": int(ws['F6'].value) if ws['F6'].value else 6,
            "nationalities": int(ws['H6'].value) if ws['H6'].value else 10,
        }
        
        message = {
            "type": "rankings_update",
            "data": {
                "institutional_data": institutional_data,
                "timestamp": asyncio.get_event_loop().time()
            }
        }
        
        # Send to all connected clients
        disconnected = set()
        for connection in active_connections:
            try:
                await connection.send_json(message)
                print(f"✅ Sent update to client")
            except Exception as e:
                print(f"❌ Error sending to client: {e}")
                disconnected.add(connection)
        
        # Remove disconnected clients
        for conn in disconnected:
            active_connections.discard(conn)
            
    except Exception as e:
        print(f"❌ Error broadcasting update: {e}")


# Global observer instance
observer = None
file_handler = None


def start_file_watcher():
    """Start watching the Excel file for changes"""
    global observer, file_handler
    
    if observer is not None:
        return  # Already watching
    
    try:
        file_handler = ExcelFileHandler(broadcast_update)
        observer = Observer()
        
        # Watch the data directory
        watch_path = os.path.dirname(EXCEL_FILE_PATH)
        observer.schedule(file_handler, watch_path, recursive=False)
        observer.start()
        
        print(f"👀 Started watching: {watch_path}")
    except Exception as e:
        print(f"❌ Error starting file watcher: {e}")


def stop_file_watcher():
    """Stop watching the Excel file"""
    global observer
    
    if observer:
        observer.stop()
        observer.join()
        observer = None
        print("🛑 Stopped file watcher")


@router.websocket("/ws/rankings")
async def rankings_websocket(websocket: WebSocket):
    """WebSocket endpoint for live rankings updates"""
    await websocket.accept()
    active_connections.add(websocket)
    
    # Start file watcher if not already running
    start_file_watcher()
    
    print(f"✅ Client connected. Total connections: {len(active_connections)}")
    
    try:
        # Send initial data
        await websocket.send_json({
            "type": "connected",
            "message": "Connected to rankings live updates"
        })
        
        # Keep connection alive
        while True:
            # Wait for messages from client (ping/pong)
            data = await websocket.receive_text()
            if data == "ping":
                await websocket.send_json({"type": "pong"})
                
    except WebSocketDisconnect:
        active_connections.discard(websocket)
        print(f"❌ Client disconnected. Remaining connections: {len(active_connections)}")
        
        # Stop watcher if no more connections
        if len(active_connections) == 0:
            stop_file_watcher()
    except Exception as e:
        print(f"❌ WebSocket error: {e}")
        active_connections.discard(websocket)
