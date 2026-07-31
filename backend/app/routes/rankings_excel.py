from fastapi import APIRouter, Depends, HTTPException
from app.auth import get_current_user
from app.models import User
import openpyxl

router = APIRouter(prefix="/api/rankings-excel", tags=["Rankings Excel"])


def check_staff_access(current_user: User):
    """Check if user is staff"""
    if current_user.account_category != "staff":
        raise HTTPException(status_code=403, detail="Only staff can access rankings")


def parse_percentage(value):
    """Parse percentage string like '~62%' or '35.1%' to float"""
    if not value:
        return 0.0
    value_str = str(value).replace('~', '').replace('%', '').strip()
    try:
        return float(value_str)
    except:
        return 0.0


@router.get("/dashboard-data")
async def get_rankings_dashboard_data(current_user: User = Depends(get_current_user)):
    """Get all rankings data from the Excel Rankings Dashboard sheet"""
    check_staff_access(current_user)
    
    try:
        # Load Excel file
        wb = openpyxl.load_workbook('/app/data/templumis_university.xlsx', data_only=True)
        ws = wb['Rankings Dashboard']
        
        # Parse institutional summary (rows 5-8)
        institutional_data = {
            "total_students": int(ws['B5'].value) if ws['B5'].value else 37,
            "ug_students": int(ws['B6'].value) if ws['B6'].value else 25,
            "pg_students": int(ws['B7'].value) if ws['B7'].value else 12,
            "faculty": int(ws['B8'].value) if ws['B8'].value else 15,
            "avg_gpa": str(ws['D5'].value) if ws['D5'].value else "3.32 / 4.0",
            "active_students": int(ws['D6'].value) if ws['D6'].value else 25,
            "graduates": int(ws['D7'].value) if ws['D7'].value else 5,
            "active_courses": int(ws['D8'].value) if ws['D8'].value else 19,
            "international_students": str(ws['F5'].value) if ws['F5'].value else "35.1%",
            "research_students": int(ws['F6'].value) if ws['F6'].value else 6,
            "student_faculty_ratio": str(ws['F7'].value) if ws['F7'].value else "2.5 : 1",
            "avg_attendance": str(ws['F8'].value) if ws['F8'].value else "83%",
            "female_ratio": str(ws['H5'].value) if ws['H5'].value else "48.6%",
            "nationalities": int(ws['H6'].value) if ws['H6'].value else 10,
            "schools_faculties": int(ws['H7'].value) if ws['H7'].value else 9,
            "avg_grade": str(ws['H8'].value) if ws['H8'].value else "77.6%",
        }
        
        # Parse Webometrics (rows 13-17)
        webometrics = {
            "name": "Webometrics Ranking of World Universities",
            "focus": "Web presence, openness & academic output",
            "overall_readiness": parse_percentage(ws['C17'].value),
            "indicators": [
                {
                    "name": str(ws['A13'].value),
                    "weight": str(ws['B13'].value),
                    "score": parse_percentage(ws['C13'].value),
                    "notes": str(ws['D13'].value),
                    "status": str(ws['E13'].value) if ws['E13'].value else ""
                },
                {
                    "name": str(ws['A14'].value),
                    "weight": str(ws['B14'].value),
                    "score": parse_percentage(ws['C14'].value),
                    "notes": str(ws['D14'].value),
                    "status": str(ws['E14'].value) if ws['E14'].value else ""
                },
                {
                    "name": str(ws['A15'].value),
                    "weight": str(ws['B15'].value),
                    "score": parse_percentage(ws['C15'].value),
                    "notes": str(ws['D15'].value),
                    "status": str(ws['E15'].value) if ws['E15'].value else ""
                },
                {
                    "name": str(ws['A16'].value),
                    "weight": str(ws['B16'].value),
                    "score": parse_percentage(ws['C16'].value),
                    "notes": str(ws['D16'].value),
                    "status": str(ws['E16'].value) if ws['E16'].value else ""
                }
            ]
        }
        
        # Parse THE (rows 22-27)
        the = {
            "name": "THE World University Rankings",
            "focus": "Teaching, research, citations & international outlook",
            "overall_readiness": parse_percentage(ws['C27'].value),
            "indicators": [
                {
                    "name": str(ws['A22'].value),
                    "weight": str(ws['B22'].value),
                    "score": parse_percentage(ws['C22'].value),
                    "notes": str(ws['D22'].value),
                    "status": str(ws['E22'].value) if ws['E22'].value else ""
                },
                {
                    "name": str(ws['A23'].value),
                    "weight": str(ws['B23'].value),
                    "score": parse_percentage(ws['C23'].value),
                    "notes": str(ws['D23'].value),
                    "status": str(ws['E23'].value) if ws['E23'].value else ""
                },
                {
                    "name": str(ws['A24'].value),
                    "weight": str(ws['B24'].value),
                    "score": parse_percentage(ws['C24'].value),
                    "notes": str(ws['D24'].value),
                    "status": str(ws['E24'].value) if ws['E24'].value else ""
                },
                {
                    "name": str(ws['A25'].value),
                    "weight": str(ws['B25'].value),
                    "score": parse_percentage(ws['C25'].value),
                    "notes": str(ws['D25'].value),
                    "status": str(ws['E25'].value) if ws['E25'].value else ""
                },
                {
                    "name": str(ws['A26'].value),
                    "weight": str(ws['B26'].value),
                    "score": parse_percentage(ws['C26'].value),
                    "notes": str(ws['D26'].value),
                    "status": str(ws['E26'].value) if ws['E26'].value else ""
                }
            ]
        }
        
        # Parse THE SSA (rows 32-37)
        the_ssa = {
            "name": "THE Sub-Saharan Africa Rankings (SSA)",
            "focus": "Regionally adapted THE criteria for African universities",
            "overall_readiness": parse_percentage(ws['C37'].value),
            "indicators": [
                {
                    "name": str(ws['A32'].value),
                    "weight": str(ws['B32'].value),
                    "score": parse_percentage(ws['C32'].value),
                    "notes": str(ws['D32'].value),
                    "status": str(ws['E32'].value) if ws['E32'].value else ""
                },
                {
                    "name": str(ws['A33'].value),
                    "weight": str(ws['B33'].value),
                    "score": parse_percentage(ws['C33'].value),
                    "notes": str(ws['D33'].value),
                    "status": str(ws['E33'].value) if ws['E33'].value else ""
                },
                {
                    "name": str(ws['A34'].value),
                    "weight": str(ws['B34'].value),
                    "score": parse_percentage(ws['C34'].value),
                    "notes": str(ws['D34'].value),
                    "status": str(ws['E34'].value) if ws['E34'].value else ""
                },
                {
                    "name": str(ws['A35'].value),
                    "weight": str(ws['B35'].value),
                    "score": parse_percentage(ws['C35'].value),
                    "notes": str(ws['D35'].value),
                    "status": str(ws['E35'].value) if ws['E35'].value else ""
                },
                {
                    "name": str(ws['A36'].value),
                    "weight": str(ws['B36'].value),
                    "score": parse_percentage(ws['C36'].value),
                    "notes": str(ws['D36'].value),
                    "status": str(ws['E36'].value) if ws['E36'].value else ""
                }
            ]
        }
        
        # Parse Shanghai ARWU (rows 42-48)
        shanghai = {
            "name": "Shanghai ARWU (Academic Ranking of World Universities)",
            "focus": "Research output, Nobel laureates & high-impact publications",
            "overall_readiness": parse_percentage(ws['C48'].value),
            "indicators": [
                {
                    "name": str(ws['A42'].value),
                    "weight": str(ws['B42'].value),
                    "score": parse_percentage(ws['C42'].value),
                    "notes": str(ws['D42'].value),
                    "status": str(ws['E42'].value) if ws['E42'].value else ""
                },
                {
                    "name": str(ws['A43'].value),
                    "weight": str(ws['B43'].value),
                    "score": parse_percentage(ws['C43'].value),
                    "notes": str(ws['D43'].value),
                    "status": str(ws['E43'].value) if ws['E43'].value else ""
                },
                {
                    "name": str(ws['A44'].value),
                    "weight": str(ws['B44'].value),
                    "score": parse_percentage(ws['C44'].value),
                    "notes": str(ws['D44'].value),
                    "status": str(ws['E44'].value) if ws['E44'].value else ""
                },
                {
                    "name": str(ws['A45'].value),
                    "weight": str(ws['B45'].value),
                    "score": parse_percentage(ws['C45'].value),
                    "notes": str(ws['D45'].value),
                    "status": str(ws['E45'].value) if ws['E45'].value else ""
                },
                {
                    "name": str(ws['A46'].value),
                    "weight": str(ws['B46'].value),
                    "score": parse_percentage(ws['C46'].value),
                    "notes": str(ws['D46'].value),
                    "status": str(ws['E46'].value) if ws['E46'].value else ""
                },
                {
                    "name": str(ws['A47'].value),
                    "weight": str(ws['B47'].value),
                    "score": parse_percentage(ws['C47'].value),
                    "notes": str(ws['D47'].value),
                    "status": str(ws['E47'].value) if ws['E47'].value else ""
                }
            ]
        }
        
        # Parse QS (rows 53-61)
        qs = {
            "name": "QS World University Rankings",
            "focus": "Reputation, faculty ratio, citations & international diversity",
            "overall_readiness": parse_percentage(ws['C61'].value),
            "indicators": [
                {
                    "name": str(ws['A53'].value),
                    "weight": str(ws['B53'].value),
                    "score": parse_percentage(ws['C53'].value),
                    "notes": str(ws['D53'].value),
                    "status": str(ws['E53'].value) if ws['E53'].value else ""
                },
                {
                    "name": str(ws['A54'].value),
                    "weight": str(ws['B54'].value),
                    "score": parse_percentage(ws['C54'].value),
                    "notes": str(ws['D54'].value),
                    "status": str(ws['E54'].value) if ws['E54'].value else ""
                },
                {
                    "name": str(ws['A55'].value),
                    "weight": str(ws['B55'].value),
                    "score": parse_percentage(ws['C55'].value),
                    "notes": str(ws['D55'].value),
                    "status": str(ws['E55'].value) if ws['E55'].value else ""
                },
                {
                    "name": str(ws['A56'].value),
                    "weight": str(ws['B56'].value),
                    "score": parse_percentage(ws['C56'].value),
                    "notes": str(ws['D56'].value),
                    "status": str(ws['E56'].value) if ws['E56'].value else ""
                },
                {
                    "name": str(ws['A57'].value),
                    "weight": str(ws['B57'].value),
                    "score": parse_percentage(ws['C57'].value),
                    "notes": str(ws['D57'].value),
                    "status": str(ws['E57'].value) if ws['E57'].value else ""
                },
                {
                    "name": str(ws['A58'].value),
                    "weight": str(ws['B58'].value),
                    "score": parse_percentage(ws['C58'].value),
                    "notes": str(ws['D58'].value),
                    "status": str(ws['E58'].value) if ws['E58'].value else ""
                },
                {
                    "name": str(ws['A59'].value),
                    "weight": str(ws['B59'].value),
                    "score": parse_percentage(ws['C59'].value),
                    "notes": str(ws['D59'].value),
                    "status": str(ws['E59'].value) if ws['E59'].value else ""
                },
                {
                    "name": str(ws['A60'].value),
                    "weight": str(ws['B60'].value),
                    "score": parse_percentage(ws['C60'].value),
                    "notes": str(ws['D60'].value),
                    "status": str(ws['E60'].value) if ws['E60'].value else ""
                }
            ]
        }
        
        # Parse CWTS Leiden (rows 66-72)
        cwts = {
            "name": "CWTS Leiden Ranking",
            "focus": "Bibliometric research performance (Web of Science)",
            "overall_readiness": parse_percentage(ws['C72'].value),
            "indicators": [
                {
                    "name": str(ws['A66'].value),
                    "weight": str(ws['B66'].value),
                    "score": parse_percentage(ws['C66'].value),
                    "notes": str(ws['D66'].value),
                    "status": str(ws['E66'].value) if ws['E66'].value else ""
                },
                {
                    "name": str(ws['A67'].value),
                    "weight": str(ws['B67'].value),
                    "score": parse_percentage(ws['C67'].value),
                    "notes": str(ws['D67'].value),
                    "status": str(ws['E67'].value) if ws['E67'].value else ""
                },
                {
                    "name": str(ws['A68'].value),
                    "weight": str(ws['B68'].value),
                    "score": parse_percentage(ws['C68'].value),
                    "notes": str(ws['D68'].value),
                    "status": str(ws['E68'].value) if ws['E68'].value else ""
                },
                {
                    "name": str(ws['A69'].value),
                    "weight": str(ws['B69'].value),
                    "score": parse_percentage(ws['C69'].value),
                    "notes": str(ws['D69'].value),
                    "status": str(ws['E69'].value) if ws['E69'].value else ""
                },
                {
                    "name": str(ws['A70'].value),
                    "weight": str(ws['B70'].value),
                    "score": parse_percentage(ws['C70'].value),
                    "notes": str(ws['D70'].value),
                    "status": str(ws['E70'].value) if ws['E70'].value else ""
                },
                {
                    "name": str(ws['A71'].value),
                    "weight": str(ws['B71'].value),
                    "score": parse_percentage(ws['C71'].value),
                    "notes": str(ws['D71'].value),
                    "status": str(ws['E71'].value) if ws['E71'].value else ""
                }
            ]
        }
        
        return {
            "institutional_data": institutional_data,
            "rankings": {
                "webometrics": webometrics,
                "the": the,
                "the_ssa": the_ssa,
                "shanghai": shanghai,
                "qs": qs,
                "cwts": cwts
            }
        }
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error reading Rankings Dashboard: {str(e)}")
