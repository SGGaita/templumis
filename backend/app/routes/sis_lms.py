from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
from fastapi.responses import FileResponse
from sqlalchemy.orm import Session
from typing import List, Optional
import openpyxl
import secrets
from pathlib import Path
from datetime import datetime

from app.excel_paths import resolve_excel_path
from app.scholarship_excel import normalize_scholarship_record
from app import scholarship_db
from app import scholarship_catalog as catalog
from app import scholarship_documents as schol_docs
from app.scholarship_applications import (
    calc_progress,
    validate_submission,
    create_reference_token,
    get_reference_by_token,
    complete_reference,
    build_alerts,
    required_field_keys,
)

from ..database import get_db
from ..auth import get_current_user, require_role
from ..models import User

router = APIRouter(prefix="/api/sis-lms", tags=["SIS/LMS"])

def _resolve_excel_path() -> Path:
    try:
        return resolve_excel_path()
    except FileNotFoundError as exc:
        raise HTTPException(status_code=404, detail=str(exc)) from exc


def load_excel_data():
    """Load data from Excel file (v2 preferred)."""
    wb = openpyxl.load_workbook(_resolve_excel_path(), read_only=True, data_only=True)
    return wb


def _load_scholarship_catalog(db: Session) -> list[dict]:
    """Published scholarship opportunities from PostgreSQL."""
    return catalog.load_programs(db, program_kind="scholarship", published_only=True)


def _open_scholarships(db: Session) -> list[dict]:
    return catalog.load_programs(db, program_kind="scholarship", published_only=True, open_only=True)


def _scholarships_lookup(db: Session) -> dict[str, dict]:
    return catalog.programs_lookup(db, program_kind="scholarship", published_only=True)


def _scholarships_lookup_list(scholarships: list[dict]) -> dict:
    return {str(s["id"]): s for s in scholarships if s.get("id")}


def _enrich_app_with_scholarship(app: dict, scholarships_dict: dict) -> dict:
    schol_id = app.get("schol_id") or app.get("scholarship_id")
    if schol_id is not None:
        app["schol_id"] = str(schol_id)
        details = scholarships_dict.get(str(schol_id))
        if details:
            app["scholarship_details"] = details
    return app


def _scholarship_apps_from_db(
    db: Session,
    student_id: str,
    scholarships_dict: dict,
    *,
    statuses_only: bool = False,
) -> list[dict]:
    """Student scholarship applications — database only (catalogue used for enrichment)."""
    return scholarship_db.applications_for_student(
        db, student_id, scholarships_dict, statuses_only=statuses_only
    )

def _normalize_key(h):
    """Normalize Excel header to snake_case key.
    'Fee ID' -> 'fee_id', 'Schol. ID' -> 'schol_id', 'Amount (KES)' -> 'amount_(kes)'
    Already-snake_case headers (student_id, etc.) are unchanged.
    """
    if h is None:
        return None
    return str(h).strip().replace(" ", "_").replace(".", "").lower()

def _row_values(sheet, row_idx: int) -> list:
    """Read a worksheet row safely (empty sheets / sparse rows won't raise)."""
    try:
        return [cell.value for cell in sheet[row_idx]]
    except (IndexError, KeyError):
        return []


def _detect_header_row(sheet, scan_rows: int = 20) -> tuple:
    """Find the header row (templumis workbooks use title rows before column headers)."""
    max_row = getattr(sheet, "max_row", None) or 0
    if max_row < 1:
        return 1, []

    header_markers = (
        "student_id",
        "course_code",
        "fee_record_id",
        "application_id",
        "payment_id",
        "grade_id",
        "schol_id",
        "scholarship_id",
        "scholarship_name",
        "grant_id",
    )

    last_row = min(scan_rows, max_row)
    for row_idx in range(1, last_row + 1):
        raw = _row_values(sheet, row_idx)
        if not any(v is not None and str(v).strip() for v in raw):
            continue
        headers = [_normalize_key(h) for h in raw]
        if "student_id" in headers:
            return row_idx, headers
        if any(h in headers for h in header_markers):
            return row_idx, headers

    raw_headers = _row_values(sheet, 1)
    return 1, [_normalize_key(h) for h in raw_headers]


def _float_field(record: dict, *keys) -> float:
    for key in keys:
        if key not in record or record[key] is None:
            continue
        try:
            return float(record[key])
        except (TypeError, ValueError):
            continue
    return 0.0


_PASSING_GRADE_PREFIXES = ("A", "B", "C", "D", "P")
_FAILING_GRADES = frozenset({"F", "W", "WF", "I", "INCOMPLETE", "NC", "U"})


def _credits_from_passing_grades(grades: list, courses_dict: dict) -> int:
    """Sum course credits for each course with a non-failing letter grade."""
    by_course: dict[str, int] = {}
    for grade in grades:
        letter = str(grade.get("letter_grade") or "").strip().upper()
        if not letter or letter in _FAILING_GRADES:
            continue
        if letter.startswith("F"):
            continue
        if not any(letter.startswith(p) for p in _PASSING_GRADE_PREFIXES):
            continue
        cc = grade.get("course_code")
        if not cc:
            continue
        credits = courses_dict.get(cc, {}).get("credits") or 0
        try:
            by_course[str(cc)] = int(float(credits))
        except (TypeError, ValueError):
            by_course[str(cc)] = 0
    return sum(by_course.values())


def _compute_credit_statistics(
    student: dict,
    enrollments: list,
    grades: list,
    courses_dict: dict,
) -> dict:
    """
    Credits earned for dashboards and scholarship forms.

    Priority for completed/earned:
    1. Students.credit_hours_earned (institutional SIS field)
    2. Enrolments with status Completed
    3. Credits from courses with passing grades
    """
    total_credits_enrolled = sum(
        e.get("course_details", {}).get("credits", 0)
        for e in enrollments
        if str(e.get("status") or "").strip().lower() == "enrolled"
    )
    completed_from_status = sum(
        e.get("course_details", {}).get("credits", 0)
        for e in enrollments
        if str(e.get("status") or "").strip().lower() == "completed"
    )
    graded_earned = _credits_from_passing_grades(grades, courses_dict)

    sheet_earned = _float_field(
        student,
        "credit_hours_earned",
        "credits_completed",
        "credit_hours",
    )

    if sheet_earned > 0:
        total_credits_completed = int(sheet_earned)
    elif completed_from_status > 0:
        total_credits_completed = int(completed_from_status)
    else:
        total_credits_completed = int(graded_earned)

    return {
        "total_credits_enrolled": int(total_credits_enrolled),
        "total_credits_completed": total_credits_completed,
        "total_credits_graded_earned": int(graded_earned),
    }


def _fee_billed(record: dict) -> float:
    return _float_field(record, "total_billed_(kes)", "total_annual", "total_billed", "sem_fee_(kes)")


def _fee_paid(record: dict) -> float:
    return _float_field(record, "amount_paid_(kes)", "total_paid", "amount_paid")


def _fee_balance(record: dict) -> float:
    return _float_field(record, "balance_(kes)", "balance_due", "balance")


def _fee_status(record: dict) -> str:
    for key in ("fee_status", "status", "payment_status"):
        val = record.get(key)
        if val is not None and str(val).strip():
            return str(val).strip()
    return "Unknown"


def sheet_to_dict_list(sheet):
    """Convert Excel sheet to list of dictionaries with normalized keys."""
    header_row_idx, headers = _detect_header_row(sheet)
    if not headers or not any(headers):
        return []
    id_col = headers.index("student_id") if "student_id" in headers else 0
    data = []
    for row in sheet.iter_rows(min_row=header_row_idx + 1, values_only=True):
        if len(row) <= id_col or row[id_col] is None:
            if id_col != 0 or row[0] is None:
                continue
        row_dict = {}
        for header, value in zip(headers, row):
            if header:
                if isinstance(value, datetime):
                    row_dict[header] = value.isoformat()
                else:
                    row_dict[header] = value
        data.append(row_dict)
    return data


def _load_student_lookups(wb):
    """Build attendance and fee lookups keyed by student_id."""
    attendance_by_student = {}
    if "Attendance" in wb.sheetnames:
        attendance_data = sheet_to_dict_list(wb["Attendance"])
        for att in attendance_data:
            student_id = att.get("student_id")
            if student_id:
                attendance_by_student.setdefault(student_id, []).append(att)

    financial_by_student = {}
    if "Fee Records" in wb.sheetnames:
        fee_records_data = sheet_to_dict_list(wb["Fee Records"])
        for fee in fee_records_data:
            student_id = fee.get("student_id")
            if student_id:
                financial_by_student.setdefault(student_id, []).append(fee)

    return attendance_by_student, financial_by_student


def _is_postgraduate(student: dict) -> bool:
    student_type = str(student.get("student_type") or "").lower()
    programme = str(student.get("programme_level") or student.get("program") or "").lower()
    return (
        "post" in student_type
        or "master" in student_type
        or "phd" in student_type
        or "doctor" in student_type
        or "mba" in programme
        or "msc" in programme
        or "mphil" in programme
        or "phd" in programme
    )


def _is_undergraduate(student: dict) -> bool:
    if _is_postgraduate(student):
        return False
    student_type = str(student.get("student_type") or "").lower()
    return "under" in student_type or "bachelor" in student_type or student_type == ""


def _compute_risk_flags(student: dict) -> list:
    """Return risk flag objects for finances, attendance, and academic progress."""
    flags = []

    compliance = student.get("compliance_status")
    avg_attendance = student.get("avg_attendance")
    if compliance == "red":
        flags.append({
            "category": "attendance",
            "severity": "critical",
            "label": "Critical attendance risk",
        })
    elif compliance == "yellow":
        flags.append({
            "category": "attendance",
            "severity": "warning",
            "label": "Attendance below target",
        })
    elif avg_attendance is not None and avg_attendance < 60:
        flags.append({
            "category": "attendance",
            "severity": "critical",
            "label": f"Attendance at {avg_attendance}%",
        })

    balance = float(student.get("fees_balance_(kes)") or student.get("balance_due") or 0)
    fees_status = str(student.get("fees_status") or student.get("payment_status") or "").lower()
    if balance > 50000:
        flags.append({
            "category": "finances",
            "severity": "critical",
            "label": f"Fee arrears KES {balance:,.0f}",
        })
    elif balance > 0 or fees_status in ("overdue", "partial", "outstanding", "defaulter", "unpaid"):
        flags.append({
            "category": "finances",
            "severity": "warning",
            "label": "Outstanding fee balance" if balance <= 0 else f"Balance KES {balance:,.0f}",
        })

    standing = str(student.get("academic_standing") or "").lower()
    gpa = student.get("gpa")
    try:
        gpa_value = float(gpa) if gpa is not None else None
    except (TypeError, ValueError):
        gpa_value = None

    if standing in ("probation", "suspended"):
        flags.append({
            "category": "academic",
            "severity": "critical",
            "label": student.get("academic_standing") or "Academic probation",
        })
    elif gpa_value is not None and gpa_value < 2.0:
        flags.append({
            "category": "academic",
            "severity": "critical",
            "label": f"GPA {gpa_value:.2f}",
        })
    elif gpa_value is not None and gpa_value < 2.5:
        flags.append({
            "category": "academic",
            "severity": "warning",
            "label": f"GPA {gpa_value:.2f}",
        })

    return flags


def _enrich_student(student: dict, attendance_by_student: dict, financial_by_student: dict) -> dict:
    """Attach compliance, financial, cohort, and risk metadata to a student row."""
    student_id = student.get("student_id")
    student_attendance = attendance_by_student.get(student_id, [])

    if student_attendance:
        attendance_values = [
            float(att.get("attendance_pct") or 0)
            for att in student_attendance
            if att.get("attendance_pct") is not None
        ]
        avg_attendance = sum(attendance_values) / len(attendance_values) if attendance_values else 0
        risk_levels = [
            str(att.get("risk_level", "")).lower()
            for att in student_attendance
            if att.get("risk_level")
        ]
        if "high" in risk_levels or avg_attendance < 60:
            student["compliance_status"] = "red"
        elif "medium" in risk_levels or avg_attendance < 75:
            student["compliance_status"] = "yellow"
        else:
            student["compliance_status"] = "green"
        student["avg_attendance"] = round(avg_attendance, 1)
    else:
        student["compliance_status"] = "green"
        student["avg_attendance"] = None

    student_fees = financial_by_student.get(student_id, [])
    excel_balance = student.get("fees_balance_(kes)")
    if student_fees:
        student["balance_due"] = sum(_fee_balance(f) for f in student_fees)
        student["total_fees"] = sum(_fee_billed(f) for f in student_fees)
        student["total_paid"] = sum(_fee_paid(f) for f in student_fees)
        student["payment_status"] = _fee_status(student_fees[0]) if student_fees else "Unknown"
    else:
        student["balance_due"] = float(excel_balance or 0)
        student["total_fees"] = 0
        student["total_paid"] = 0
        student["payment_status"] = student.get("fees_status", "Unknown")

    if excel_balance is not None and not student_fees:
        try:
            student["balance_due"] = float(excel_balance)
        except (TypeError, ValueError):
            pass

    student["cohort_level"] = "postgraduate" if _is_postgraduate(student) else "undergraduate"
    student["risk_flags"] = _compute_risk_flags(student)
    student["is_at_risk"] = len(student["risk_flags"]) > 0
    student["risk_categories"] = sorted({f["category"] for f in student["risk_flags"]})
    return student


def _load_enriched_students():
    wb = load_excel_data()
    students = sheet_to_dict_list(wb["Students"])
    attendance_by_student, financial_by_student = _load_student_lookups(wb)
    wb.close()
    return [_enrich_student(s, attendance_by_student, financial_by_student) for s in students]


@router.get("/students")
async def get_students(
    search: Optional[str] = None,
    program: Optional[str] = None,
    year_of_study: Optional[str] = None,
    status: Optional[str] = None,
    cohort_level: Optional[str] = Query(None, description="undergraduate or postgraduate"),
    at_risk_only: bool = False,
    skip: int = 0,
    limit: int = 100,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Get students from SIS"""
    require_role(current_user, ["staff", "global_admin"])

    students = _load_enriched_students()

    if search:
        search_lower = search.lower()
        students = [
            s for s in students
            if search_lower in str(s.get("full_name", "")).lower()
            or search_lower in str(s.get("student_id", "")).lower()
            or search_lower in str(s.get("email", "")).lower()
        ]

    if program:
        students = [s for s in students if s.get("program") == program]

    if year_of_study:
        students = [s for s in students if s.get("year_of_study") == year_of_study]

    if status:
        students = [s for s in students if str(s.get("status", "")).lower() == status.lower()]

    if cohort_level:
        level = cohort_level.lower()
        if level in ("undergraduate", "undergrad", "ug"):
            students = [s for s in students if s.get("cohort_level") == "undergraduate"]
        elif level in ("postgraduate", "postgrad", "pg"):
            students = [s for s in students if s.get("cohort_level") == "postgraduate"]

    if at_risk_only:
        students = [s for s in students if s.get("is_at_risk")]

    total = len(students)
    students = students[skip:skip + limit]

    return {
        "students": students,
        "total": total,
        "skip": skip,
        "limit": limit
    }


@router.get("/at-risk/summary")
async def get_at_risk_summary(
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Counts of at-risk students by category and cohort."""
    require_role(current_user, ["staff", "global_admin"])

    students = [s for s in _load_enriched_students() if s.get("is_at_risk")]
    summary = {
        "total": len(students),
        "undergraduate": len([s for s in students if s.get("cohort_level") == "undergraduate"]),
        "postgraduate": len([s for s in students if s.get("cohort_level") == "postgraduate"]),
        "by_category": {
            "finances": 0,
            "attendance": 0,
            "academic": 0,
        },
    }
    for student in students:
        for category in student.get("risk_categories", []):
            if category in summary["by_category"]:
                summary["by_category"][category] += 1
    return summary


@router.get("/at-risk")
async def get_at_risk_students(
    category: Optional[str] = Query(None, description="finances, attendance, or academic"),
    cohort_level: Optional[str] = Query(None, description="undergraduate or postgraduate"),
    search: Optional[str] = None,
    skip: int = 0,
    limit: int = 200,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Students flagged at risk by finances, attendance, or academic progress."""
    require_role(current_user, ["staff", "global_admin"])

    students = [s for s in _load_enriched_students() if s.get("is_at_risk")]

    if category:
        cat = category.lower()
        students = [s for s in students if cat in s.get("risk_categories", [])]

    if cohort_level:
        level = cohort_level.lower()
        if level in ("undergraduate", "undergrad", "ug"):
            students = [s for s in students if s.get("cohort_level") == "undergraduate"]
        elif level in ("postgraduate", "postgrad", "pg"):
            students = [s for s in students if s.get("cohort_level") == "postgraduate"]

    if search:
        search_lower = search.lower()
        students = [
            s for s in students
            if search_lower in str(s.get("full_name", "")).lower()
            or search_lower in str(s.get("student_id", "")).lower()
            or search_lower in str(s.get("email", "")).lower()
        ]

    total = len(students)
    page = students[skip:skip + limit]
    return {
        "students": page,
        "total": total,
        "skip": skip,
        "limit": limit,
        "summary": {
            "finances": len([s for s in students if "finances" in s.get("risk_categories", [])]),
            "attendance": len([s for s in students if "attendance" in s.get("risk_categories", [])]),
            "academic": len([s for s in students if "academic" in s.get("risk_categories", [])]),
        },
    }

@router.get("/my-profile")
async def get_my_profile(
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Get the logged-in student's full profile from Excel"""
    student_id = current_user.student_registration_number
    if not student_id:
        raise HTTPException(status_code=400, detail="No student registration number linked to this account")

    wb = load_excel_data()

    students_sheet = wb["Students"]
    students = sheet_to_dict_list(students_sheet)
    student = next((s for s in students if s.get("student_id") == student_id), None)
    if not student:
        wb.close()
        raise HTTPException(status_code=404, detail="Student record not found in institutional data")

    enrollments_sheet = wb["Enrolments"]
    all_enrollments = sheet_to_dict_list(enrollments_sheet)
    enrollments = [e for e in all_enrollments if e.get("student_id") == student_id]

    grades_sheet = wb["Grades"]
    all_grades = sheet_to_dict_list(grades_sheet)
    grades = [g for g in all_grades if g.get("student_id") == student_id]

    attendance = []
    if "Attendance" in wb.sheetnames:
        attendance_sheet = wb["Attendance"]
        all_attendance = sheet_to_dict_list(attendance_sheet)
        attendance = [a for a in all_attendance if a.get("student_id") == student_id]

    fee_records = []
    if "Fee Records" in wb.sheetnames:
        fee_records_sheet = wb["Fee Records"]
        all_fee_records = sheet_to_dict_list(fee_records_sheet)
        fee_records = [f for f in all_fee_records if f.get("student_id") == student_id]

    payments = []
    if "Payments" in wb.sheetnames:
        payments_sheet = wb["Payments"]
        all_payments = sheet_to_dict_list(payments_sheet)
        payments = [p for p in all_payments if p.get("student_id") == student_id]

    scholarships_dict = _scholarships_lookup(db)
    scholarship_apps = _scholarship_apps_from_db(db, student_id, scholarships_dict)

    courses_sheet = wb["Courses"]
    all_courses = sheet_to_dict_list(courses_sheet)
    courses_dict = {c["course_code"]: c for c in all_courses if c.get("course_code")}

    for enrollment in enrollments:
        course_code = enrollment.get("course_code")
        if course_code and course_code in courses_dict:
            enrollment["course_details"] = courses_dict[course_code]

    for grade in grades:
        course_code = grade.get("course_code")
        if course_code and course_code in courses_dict:
            grade["course_details"] = courses_dict[course_code]

    for att in attendance:
        course_code = att.get("course_code")
        if course_code and course_code in courses_dict:
            att["course_details"] = courses_dict[course_code]

    wb.close()

    # Use pre-calculated GPA from Students sheet; fall back to letter-grade computation
    excel_gpa = student.get("gpa")
    if excel_gpa is not None:
        try:
            gpa = round(float(excel_gpa), 2)
        except (ValueError, TypeError):
            excel_gpa = None
    if excel_gpa is None:
        grade_points = {"A": 4.0, "B": 3.0, "C": 2.0, "D": 1.0, "F": 0.0}
        total_points = 0
        total_credits_graded = 0
        for grade in grades:
            letter = grade.get("letter_grade")
            if letter and letter in grade_points:
                credits = courses_dict.get(grade.get("course_code"), {}).get("credits", 0)
                total_points += grade_points[letter] * credits
                total_credits_graded += credits
        gpa = round(total_points / total_credits_graded, 2) if total_credits_graded > 0 else 0.0

    total_sessions = sum(a.get("total_sessions", 0) for a in attendance)
    total_present = sum(a.get("present", 0) for a in attendance)
    attendance_rate = round((total_present / total_sessions * 100), 1) if total_sessions > 0 else 0

    from app.awards import AWARD_CREDITED, credited_total_for_student

    total_fees = sum(_fee_billed(f) for f in fee_records)
    total_paid = sum(_fee_paid(f) for f in fee_records)
    balance_due = sum(_fee_balance(f) for f in fee_records)
    tuition_credits = credited_total_for_student(db, student_id)
    legacy_scholarships = sum(
        _float_field(a, "award_amount_(kes)", "award_amount")
        for a in scholarship_apps
        if str(a.get("status", "")).lower() in ("awarded", "approved")
        and str(a.get("award_stage") or "") != AWARD_CREDITED
    )
    total_scholarships = tuition_credits + legacy_scholarships
    net_balance_due = max(0.0, balance_due - tuition_credits)
    credit_stats = _compute_credit_statistics(student, enrollments, grades, courses_dict)
    student = {
        **student,
        "credits_completed": credit_stats["total_credits_completed"],
        "cohort_level": "postgraduate" if _is_postgraduate(student) else "undergraduate",
    }

    return {
        "student": student,
        "enrollments": enrollments,
        "grades": grades,
        "attendance": attendance,
        "fee_records": fee_records,
        "payments": payments,
        "scholarship_apps": scholarship_apps,
        "statistics": {
            "gpa": gpa,
            "total_credits_enrolled": credit_stats["total_credits_enrolled"],
            "total_credits_completed": credit_stats["total_credits_completed"],
            "total_credits_graded_earned": credit_stats["total_credits_graded_earned"],
            "total_courses_enrolled": len([e for e in enrollments if e.get("status") == "Enrolled"]),
            "total_courses_completed": len([e for e in enrollments if e.get("status") == "Completed"]),
            "attendance_rate": attendance_rate,
            "total_fees": total_fees,
            "total_paid": total_paid,
            "balance_due": balance_due,
            "tuition_scholarship_credits": tuition_credits,
            "net_balance_due": net_balance_due,
            "total_payments": len(payments),
            "total_scholarships": total_scholarships,
            "approved_scholarships": len([
                a for a in scholarship_apps
                if a.get("award_stage") == AWARD_CREDITED
                or str(a.get("status", "")).lower() in ("approved", "awarded")
            ]),
            "total_sessions": total_sessions,
            "total_present": total_present
        }
    }


@router.get("/students/{student_id}")
async def get_student_detail(
    student_id: str,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Get detailed student information including enrollments, grades, attendance, and financial data"""
    require_role(current_user, ["staff", "global_admin"])
    
    wb = load_excel_data()
    
    # Get student info
    students_sheet = wb["Students"]
    students = sheet_to_dict_list(students_sheet)
    student = next((s for s in students if s.get("student_id") == student_id), None)
    
    if not student:
        wb.close()
        raise HTTPException(status_code=404, detail="Student not found")
    
    # Get enrollments
    enrollments_sheet = wb["Enrolments"]
    all_enrollments = sheet_to_dict_list(enrollments_sheet)
    enrollments = [e for e in all_enrollments if e.get("student_id") == student_id]
    
    # Get grades
    grades_sheet = wb["Grades"]
    all_grades = sheet_to_dict_list(grades_sheet)
    grades = [g for g in all_grades if g.get("student_id") == student_id]
    
    # Get attendance
    attendance = []
    if "Attendance" in wb.sheetnames:
        attendance_sheet = wb["Attendance"]
        all_attendance = sheet_to_dict_list(attendance_sheet)
        attendance = [a for a in all_attendance if a.get("student_id") == student_id]
    
    # Get financial data
    fee_records = []
    if "Fee Records" in wb.sheetnames:
        fee_records_sheet = wb["Fee Records"]
        all_fee_records = sheet_to_dict_list(fee_records_sheet)
        fee_records = [f for f in all_fee_records if f.get("student_id") == student_id]
    
    payments = []
    if "Payments" in wb.sheetnames:
        payments_sheet = wb["Payments"]
        all_payments = sheet_to_dict_list(payments_sheet)
        payments = [p for p in all_payments if p.get("student_id") == student_id]
    
    scholarships_dict = _scholarships_lookup(db)
    scholarship_apps = _scholarship_apps_from_db(db, student_id, scholarships_dict)

    # Get courses for enrichment
    courses_sheet = wb["Courses"]
    all_courses = sheet_to_dict_list(courses_sheet)
    courses_dict = {c["course_code"]: c for c in all_courses}
    
    # Enrich enrollments with course details
    for enrollment in enrollments:
        course_code = enrollment.get("course_code")
        if course_code and course_code in courses_dict:
            enrollment["course_details"] = courses_dict[course_code]
    
    # Enrich grades with course details
    for grade in grades:
        course_code = grade.get("course_code")
        if course_code and course_code in courses_dict:
            grade["course_details"] = courses_dict[course_code]
    
    # Enrich attendance with course details
    for att in attendance:
        course_code = att.get("course_code")
        if course_code and course_code in courses_dict:
            att["course_details"] = courses_dict[course_code]
    
    wb.close()
    
    credit_stats = _compute_credit_statistics(student, enrollments, grades, courses_dict)
    completed_courses = [e for e in enrollments if str(e.get("status") or "").lower() == "completed"]
    student = {**student, "credits_completed": credit_stats["total_credits_completed"]}
    
    # Use pre-calculated GPA from Students sheet; fall back to letter-grade computation
    excel_gpa = student.get("gpa")
    if excel_gpa is not None:
        try:
            gpa = round(float(excel_gpa), 2)
        except (ValueError, TypeError):
            excel_gpa = None
    if excel_gpa is None:
        grade_points = {"A": 4.0, "B": 3.0, "C": 2.0, "D": 1.0, "F": 0.0}
        total_points = 0
        total_credits_graded = 0
        for grade in grades:
            letter = grade.get("letter_grade")
            if letter and letter in grade_points:
                course_code = grade.get("course_code")
                course = courses_dict.get(course_code, {})
                credits = course.get("credits", 0)
                total_points += grade_points[letter] * credits
                total_credits_graded += credits
        gpa = round(total_points / total_credits_graded, 2) if total_credits_graded > 0 else 0.0
    
    # Calculate attendance stats
    total_sessions = sum(a.get("total_sessions", 0) for a in attendance)
    total_present = sum(a.get("present", 0) for a in attendance)
    attendance_rate = round((total_present / total_sessions * 100), 1) if total_sessions > 0 else 0
    
    # Calculate financial statistics
    total_fees = sum(_fee_billed(f) for f in fee_records)
    total_paid = sum(_fee_paid(f) for f in fee_records)
    balance_due = sum(_fee_balance(f) for f in fee_records)
    total_payments = sum(_float_field(p, "amount_(kes)", "amount") for p in payments)
    total_scholarships = sum(_fee_billed(f) for f in fee_records)
    
    # Get approved scholarships
    approved_scholarships = [
        app for app in scholarship_apps
        if str(app.get("status", "")).lower() in ("approved", "awarded")
    ]
    
    return {
        "student": student,
        "enrollments": enrollments,
        "grades": grades,
        "attendance": attendance,
        "fee_records": fee_records,
        "payments": payments,
        "scholarship_apps": scholarship_apps,
        "statistics": {
            "gpa": gpa,
            "total_credits_enrolled": credit_stats["total_credits_enrolled"],
            "total_credits_completed": credit_stats["total_credits_completed"],
            "total_credits_graded_earned": credit_stats["total_credits_graded_earned"],
            "total_courses_enrolled": len([e for e in enrollments if str(e.get("status") or "").lower() == "enrolled"]),
            "total_courses_completed": len(completed_courses),
            "attendance_rate": attendance_rate,
            "total_fees": total_fees,
            "total_paid": total_paid,
            "balance_due": balance_due,
            "total_payments": len(payments),
            "total_scholarships": total_scholarships,
            "approved_scholarships": len(approved_scholarships),
            "total_sessions": total_sessions,
            "total_present": total_present
        }
    }

@router.get("/courses")
async def get_courses(
    search: Optional[str] = None,
    department: Optional[str] = None,
    status: Optional[str] = None,
    skip: int = 0,
    limit: int = 100,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Get courses from LMS"""
    require_role(current_user, ["staff", "global_admin"])
    
    wb = load_excel_data()
    courses_sheet = wb["Courses"]
    courses = sheet_to_dict_list(courses_sheet)
    wb.close()
    
    # Apply filters
    if search:
        search_lower = search.lower()
        courses = [c for c in courses if 
                  search_lower in c.get("title", "").lower() or 
                  search_lower in c.get("course_code", "").lower()]
    
    if department:
        courses = [c for c in courses if c.get("department") == department]
    
    if status:
        courses = [c for c in courses if c.get("status") == status]
    
    total = len(courses)
    courses = courses[skip:skip + limit]
    
    return {
        "courses": courses,
        "total": total,
        "skip": skip,
        "limit": limit
    }

@router.get("/enrollments")
async def get_enrollments(
    student_id: Optional[str] = None,
    course_code: Optional[str] = None,
    status: Optional[str] = None,
    skip: int = 0,
    limit: int = 100,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Get enrollments"""
    require_role(current_user, ["staff", "global_admin"])
    
    wb = load_excel_data()
    enrollments_sheet = wb["Enrolments"]
    enrollments = sheet_to_dict_list(enrollments_sheet)
    wb.close()
    
    # Apply filters
    if student_id:
        enrollments = [e for e in enrollments if e.get("student_id") == student_id]
    
    if course_code:
        enrollments = [e for e in enrollments if e.get("course_code") == course_code]
    
    if status:
        enrollments = [e for e in enrollments if e.get("status") == status]
    
    total = len(enrollments)
    enrollments = enrollments[skip:skip + limit]
    
    return {
        "enrollments": enrollments,
        "total": total,
        "skip": skip,
        "limit": limit
    }

@router.get("/stats")
async def get_stats(
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Get overall statistics"""
    require_role(current_user, ["staff", "global_admin"])
    
    wb = load_excel_data()
    
    students_sheet = wb["Students"]
    students = sheet_to_dict_list(students_sheet)
    
    courses_sheet = wb["Courses"]
    courses = sheet_to_dict_list(courses_sheet)
    
    enrollments_sheet = wb["Enrolments"]
    enrollments = sheet_to_dict_list(enrollments_sheet)
    
    students_by_cohort = {"undergraduate": 0, "postgraduate": 0}
    for student in students:
        if _is_postgraduate(student):
            students_by_cohort["postgraduate"] += 1
        else:
            students_by_cohort["undergraduate"] += 1

    # Calculate stats
    total_students = len(students)
    active_students = len([s for s in students if s.get("status") == "Active"])
    total_courses = len(courses)
    active_courses = len([c for c in courses if c.get("status") == "Active"])
    total_enrollments = len(enrollments)
    
    # Students by program
    programs = {}
    for student in students:
        prog = student.get("program", "Unknown")
        programs[prog] = programs.get(prog, 0) + 1
    
    # Students by year
    years = {}
    for student in students:
        year = student.get("year_of_study", "Unknown")
        years[year] = years.get(year, 0) + 1
    
    # Students by gender
    genders = {}
    for student in students:
        gender = student.get("gender", "Unknown")
        genders[gender] = genders.get(gender, 0) + 1
    
    # Students by nationality
    nationalities = {}
    for student in students:
        nationality = student.get("nationality", "Unknown")
        nationalities[nationality] = nationalities.get(nationality, 0) + 1
    
    # Students by major
    majors = {}
    for student in students:
        major = student.get("major", "Unknown")
        majors[major] = majors.get(major, 0) + 1
    
    # Gender breakdown by major
    gender_by_major = {}
    for student in students:
        major = student.get("major", "Unknown")
        gender = student.get("gender", "Unknown")
        if major not in gender_by_major:
            gender_by_major[major] = {"Male": 0, "Female": 0, "Other": 0}
        if gender in gender_by_major[major]:
            gender_by_major[major][gender] += 1
        else:
            gender_by_major[major]["Other"] += 1
    
    # Nationality breakdown by major
    nationality_by_major = {}
    for student in students:
        major = student.get("major", "Unknown")
        nationality = student.get("nationality", "Unknown")
        if major not in nationality_by_major:
            nationality_by_major[major] = {}
        nationality_by_major[major][nationality] = nationality_by_major[major].get(nationality, 0) + 1
    
    # Students by enrollment date (cohorts)
    cohorts = {}
    for student in students:
        enroll_date = student.get("enrollment_date", "")
        if enroll_date:
            # Extract year from enrollment date
            try:
                if isinstance(enroll_date, str):
                    year = enroll_date.split("-")[0] if "-" in enroll_date else enroll_date[:4]
                else:
                    year = str(enroll_date.year)
                cohorts[f"Cohort {year}"] = cohorts.get(f"Cohort {year}", 0) + 1
            except:
                cohorts["Unknown"] = cohorts.get("Unknown", 0) + 1
    
    # Calculate compliance statistics from attendance
    attendance_sheet = wb["Attendance"]
    attendance_data = sheet_to_dict_list(attendance_sheet)
    
    # Create attendance lookup by student_id
    attendance_by_student = {}
    for att in attendance_data:
        student_id = att.get("student_id")
        if student_id:
            if student_id not in attendance_by_student:
                attendance_by_student[student_id] = []
            attendance_by_student[student_id].append(att)
    
    # Calculate compliance for each student
    compliance_stats = {"green": 0, "yellow": 0, "red": 0}
    for student in students:
        student_id = student.get("student_id")
        student_attendance = attendance_by_student.get(student_id, [])
        
        if student_attendance:
            # Filter out None values and convert to float
            attendance_values = [float(att.get("attendance_pct") or 0) for att in student_attendance if att.get("attendance_pct") is not None]
            
            if attendance_values:
                avg_attendance = sum(attendance_values) / len(attendance_values)
            else:
                avg_attendance = 0
            
            risk_levels = [str(att.get("risk_level", "")).lower() for att in student_attendance if att.get("risk_level")]
            
            if "high" in risk_levels or avg_attendance < 60:
                compliance_stats["red"] += 1
            elif "medium" in risk_levels or avg_attendance < 75:
                compliance_stats["yellow"] += 1
            else:
                compliance_stats["green"] += 1
        else:
            compliance_stats["green"] += 1
    
    # Get financial statistics (if sheets exist)
    fee_records = []
    payments = []
    
    try:
        if "Fee Records" in wb.sheetnames:
            fee_records_sheet = wb["Fee Records"]
            fee_records = sheet_to_dict_list(fee_records_sheet)
    except:
        pass
    
    try:
        if "Payments" in wb.sheetnames:
            payments_sheet = wb["Payments"]
            payments = sheet_to_dict_list(payments_sheet)
    except:
        pass
    
    # Calculate financial totals
    total_fees_due = sum(_fee_billed(f) for f in fee_records) if fee_records else 0
    total_fees_paid = sum(_fee_paid(f) for f in fee_records) if fee_records else 0
    total_balance_due = sum(_fee_balance(f) for f in fee_records) if fee_records else 0

    app_stats = scholarship_db.institution_application_stats(db)

    # Payment status breakdown (from fee records)
    payment_status = {}
    for record in fee_records:
        status = _fee_status(record)
        payment_status[status] = payment_status.get(status, 0) + 1

    # Available scholarships by type (catalogue from database)
    scholarship_types = {}
    scholarships = catalog.load_programs(db, program_kind="scholarship", published_only=True)
    for schol in scholarships:
        schol_type = schol.get("type", "Unknown")
        scholarship_types[schol_type] = scholarship_types.get(schol_type, 0) + 1
    
    wb.close()
    
    return {
        "total_students": total_students,
        "active_students": active_students,
        "total_courses": total_courses,
        "active_courses": active_courses,
        "total_enrollments": total_enrollments,
        "students_by_program": programs,
        "students_by_year": years,
        "students_by_gender": genders,
        "students_by_nationality": nationalities,
        "students_by_major": majors,
        "gender_by_major": gender_by_major,
        "nationality_by_major": nationality_by_major,
        "students_by_cohort": cohorts,
        "students_by_level": students_by_cohort,
        "compliance_status": compliance_stats,
        "financial": {
            "total_fees_due": total_fees_due,
            "total_fees_paid": total_fees_paid,
            "total_balance_due": total_balance_due,
            "total_scholarships_awarded": app_stats["total_scholarships_awarded"],
            "payment_status": payment_status,
            "scholarship_application_status": app_stats["scholarship_application_status"],
            "scholarship_types": scholarship_types,
            "total_payments": len(payments),
            "total_scholarship_apps": app_stats["total_scholarship_apps"],
            "total_scholarships_available": len(scholarships),
        }
    }

def _parse_pct_value(value) -> float:
    if value is None:
        return 0.0
    text = str(value).replace("~", "").replace("%", "").strip()
    try:
        return float(text)
    except (TypeError, ValueError):
        return 0.0


def _compute_leadership_extras(students: list, courses: list, enrollments: list) -> dict:
    """Additional leadership metrics from SIS/LMS records."""
    total = len(students) or 1
    gpas = []
    for s in students:
        try:
            if s.get("gpa") is not None:
                gpas.append(float(s.get("gpa")))
        except (TypeError, ValueError):
            pass

    gpa_avg = round(sum(gpas) / len(gpas), 2) if gpas else 0
    gpa_bands = [
        {"name": "Excellent (3.5+)", "count": sum(1 for g in gpas if g >= 3.5), "color": "green"},
        {"name": "Satisfactory (2.5–3.49)", "count": sum(1 for g in gpas if 2.5 <= g < 3.5), "color": "blue"},
        {"name": "At risk (<2.5)", "count": sum(1 for g in gpas if g < 2.5), "color": "red"},
    ]

    standing_counts = {}
    fees_counts = {}
    clearance_counts = {}
    departments = {}
    for s in students:
        standing = str(s.get("academic_standing") or "Unknown")
        standing_counts[standing] = standing_counts.get(standing, 0) + 1
        fees = str(s.get("fees_status") or "Unknown")
        fees_counts[fees] = fees_counts.get(fees, 0) + 1
        clearance = str(s.get("graduation_clearance") or "Not Yet")
        clearance_counts[clearance] = clearance_counts.get(clearance, 0) + 1
        dept = str(s.get("department") or "Unknown")
        departments[dept] = departments.get(dept, 0) + 1

    departments_top = sorted(
        [{"name": k, "count": v, "share_pct": round(v / total * 100, 1)} for k, v in departments.items()],
        key=lambda x: x["count"],
        reverse=True,
    )[:6]

    nationalities = {}
    genders = {"Male": 0, "Female": 0, "Other": 0}
    international = 0
    for s in students:
        nat = str(s.get("nationality") or "Unknown")
        nationalities[nat] = nationalities.get(nat, 0) + 1
        if nat.lower() not in ("kenyan", "kenya", "local"):
            international += 1
        gender = str(s.get("gender") or "Other")
        if gender in genders:
            genders[gender] += 1
        else:
            genders["Other"] += 1

    female_pct = round((genders["Female"] / total) * 100, 1)
    international_pct = round((international / total) * 100, 1)

    active_courses = [c for c in courses if str(c.get("status", "")).lower() == "active"]
    total_course_capacity = sum(int(c.get("enrolled_count") or 0) for c in courses)
    avg_class_size = round(total_course_capacity / len(active_courses), 1) if active_courses else 0
    students_per_course = round(len(enrollments) / len(active_courses), 1) if active_courses else 0

    probation_count = standing_counts.get("Probation", 0) + standing_counts.get("Suspended", 0)
    cleared_grad = clearance_counts.get("Cleared", 0)
    near_graduation = sum(
        1 for s in students
        if "graduat" in str(s.get("journey_stage", "")).lower()
        or "year 4" in str(s.get("year_of_study", "")).lower()
        or "year 5" in str(s.get("year_of_study", "")).lower()
    )

    major_counts = {}
    for s in students:
        m = str(s.get("major") or "Unknown")
        major_counts[m] = major_counts.get(m, 0) + 1
    shares = sorted([(c / total) for c in major_counts.values()], reverse=True)
    concentration_index = round(sum(s * s for s in shares) * 100, 1) if shares else 0

    return {
        "academic_standing": standing_counts,
        "fees_status": fees_counts,
        "graduation_clearance": clearance_counts,
        "gpa": {"average": gpa_avg, "bands": gpa_bands},
        "departments_top": departments_top,
        "internationalization": {
            "international_pct": international_pct,
            "nationality_count": len(nationalities),
            "female_pct": female_pct,
        },
        "operational": {
            "active_courses": len(active_courses),
            "total_courses": len(courses),
            "total_enrollments": len(enrollments),
            "avg_class_size": avg_class_size,
            "students_per_active_course": students_per_course,
        },
        "student_success": {
            "probation_or_suspended": probation_count,
            "near_graduation": near_graduation,
            "graduation_cleared": cleared_grad,
        },
        "program_concentration_index": concentration_index,
    }


def _parse_rankings_snapshot(wb) -> dict:
    """Lightweight rankings readiness from Excel dashboard sheet."""
    if "Rankings Dashboard" not in wb.sheetnames:
        return None
    try:
        ws = wb["Rankings Dashboard"]

        def overall_at_row(row: int) -> float:
            return _parse_pct_value(ws[f"C{row}"].value)

        systems = [
            {"id": "webometrics", "name": "Webometrics", "readiness_pct": overall_at_row(17)},
            {"id": "the", "name": "THE World", "readiness_pct": overall_at_row(27)},
            {"id": "qs", "name": "QS World", "readiness_pct": overall_at_row(61)},
        ]
        systems = [s for s in systems if s["readiness_pct"] > 0]
        avg_readiness = round(sum(s["readiness_pct"] for s in systems) / len(systems), 1) if systems else 0

        return {
            "institutional_profile": {
                "faculty_count": int(ws["B8"].value or 0),
                "student_faculty_ratio": str(ws["F7"].value or "—"),
                "avg_attendance": str(ws["F8"].value or "—"),
                "international_students": str(ws["F5"].value or "—"),
                "female_ratio": str(ws["H5"].value or "—"),
                "schools_faculties": int(ws["H7"].value or 0),
            },
            "ranking_systems": systems,
            "avg_ranking_readiness_pct": avg_readiness,
        }
    except Exception:
        return None


@router.get("/analytics/executive")
async def get_executive_analytics(
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Condensed analytics payload for leadership dashboards."""
    require_role(current_user, ["staff", "global_admin"])

    wb = load_excel_data()
    students = sheet_to_dict_list(wb["Students"])
    courses = sheet_to_dict_list(wb["Courses"])
    enrollments = sheet_to_dict_list(wb["Enrolments"])
    leadership = _compute_leadership_extras(students, courses, enrollments)
    rankings_snapshot = _parse_rankings_snapshot(wb)
    wb.close()

    stats = await get_stats(current_user=current_user, db=db)
    at_risk_students = [s for s in _load_enriched_students() if s.get("is_at_risk")]

    total = stats.get("total_students") or 0
    majors_sorted = sorted(
        [{"name": k, "count": v} for k, v in (stats.get("students_by_major") or {}).items()],
        key=lambda x: x["count"],
        reverse=True,
    )
    for item in majors_sorted:
        item["share_pct"] = round((item["count"] / total * 100), 1) if total else 0

    nationalities_ranked = sorted(
        [{"name": k, "count": v} for k, v in (stats.get("students_by_nationality") or {}).items()],
        key=lambda x: x["count"],
        reverse=True,
    )
    domestic_names = {"kenyan", "kenya", "local"}
    for item in nationalities_ranked:
        item["share_pct"] = round((item["count"] / total * 100), 1) if total else 0
        item["is_domestic"] = item["name"].lower().strip() in domestic_names
    nationalities_top = nationalities_ranked[:6]

    compliance = stats.get("compliance_status") or {}
    compliance_total = sum(compliance.values()) or 1
    at_risk_count = len(at_risk_students)
    financial = stats.get("financial") or {}
    fees_due = float(financial.get("total_fees_due") or 0)
    fees_paid = float(financial.get("total_fees_paid") or 0)
    collection_rate = round((fees_paid / fees_due * 100), 1) if fees_due > 0 else 0

    insights = []
    if majors_sorted:
        top = majors_sorted[0]
        insights.append({
            "priority": "high",
            "title": "Enrollment concentration",
            "message": f"{top['name']} is the largest major ({top['share_pct']}% of students). Review capacity and faculty allocation.",
        })
    if len(majors_sorted) >= 2:
        bottom = majors_sorted[-1]
        insights.append({
            "priority": "medium",
            "title": "Smallest programme",
            "message": f"{bottom['name']} has the fewest students ({bottom['count']}). Consider recruitment or consolidation options.",
        })
    if at_risk_count > 0 and total > 0:
        pct = round(at_risk_count / total * 100, 1)
        insights.append({
            "priority": "high" if pct >= 10 else "medium",
            "title": "Student success alert",
            "message": f"{at_risk_count} students ({pct}%) are flagged across finances, attendance, or academics.",
        })
    if collection_rate < 85 and fees_due > 0:
        insights.append({
            "priority": "high",
            "title": "Revenue collection",
            "message": f"Fee collection is at {collection_rate}%. Outstanding balance requires executive follow-up.",
        })
    elif collection_rate >= 90:
        insights.append({
            "priority": "low",
            "title": "Strong collections",
            "message": f"Fee collection at {collection_rate}% is healthy relative to billed amounts.",
        })

    intl = leadership["internationalization"]
    if intl["international_pct"] >= 30:
        insights.append({
            "priority": "medium",
            "title": "Global profile",
            "message": f"{intl['international_pct']}% international students across {intl['nationality_count']} nationalities — leverage for rankings and partnerships.",
        })

    probation = leadership["student_success"]["probation_or_suspended"]
    if probation > 0:
        insights.append({
            "priority": "high",
            "title": "Academic standing",
            "message": f"{probation} students on probation or suspended — academic board review recommended.",
        })

    if leadership["program_concentration_index"] >= 25:
        insights.append({
            "priority": "medium",
            "title": "Program concentration",
            "message": f"Enrollment concentration index is {leadership['program_concentration_index']} — portfolio may be over-dependent on a few majors.",
        })

    if rankings_snapshot and rankings_snapshot.get("avg_ranking_readiness_pct", 0) < 60:
        insights.append({
            "priority": "high",
            "title": "Rankings readiness",
            "message": f"Average ranking-system readiness is {rankings_snapshot['avg_ranking_readiness_pct']}%. Prioritize data and research visibility investments.",
        })

    benchmarks = [
        {
            "label": "Fee collection",
            "value": collection_rate,
            "target": 90,
            "unit": "%",
            "status": "good" if collection_rate >= 90 else "watch" if collection_rate >= 75 else "critical",
        },
        {
            "label": "Students on track",
            "value": round((compliance.get("green", 0) / compliance_total) * 100, 1),
            "target": 85,
            "unit": "%",
            "status": "good" if compliance.get("green", 0) / compliance_total >= 0.85 else "watch",
        },
        {
            "label": "At-risk share",
            "value": round(at_risk_count / total * 100, 1) if total else 0,
            "target": 8,
            "unit": "%",
            "status": (
                "good" if total and at_risk_count / total <= 0.08
                else "watch" if total and at_risk_count / total <= 0.15
                else "critical"
            ),
            "lower_is_better": True,
        },
        {
            "label": "Institution GPA",
            "value": leadership["gpa"]["average"],
            "target": 3.0,
            "unit": "",
            "status": "good" if leadership["gpa"]["average"] >= 3.0 else "watch",
        },
        {
            "label": "International students",
            "value": intl["international_pct"],
            "target": 25,
            "unit": "%",
            "status": "good" if intl["international_pct"] >= 25 else "watch",
        },
    ]
    if rankings_snapshot:
        benchmarks.append({
            "label": "Ranking readiness",
            "value": rankings_snapshot["avg_ranking_readiness_pct"],
            "target": 70,
            "unit": "%",
            "status": "good" if rankings_snapshot["avg_ranking_readiness_pct"] >= 70 else "watch" if rankings_snapshot["avg_ranking_readiness_pct"] >= 50 else "critical",
        })

    risk_by_category = {"finances": 0, "attendance": 0, "academic": 0}
    for student in at_risk_students:
        for cat in student.get("risk_categories", []):
            if cat in risk_by_category:
                risk_by_category[cat] += 1

    return {
        "generated_at": datetime.utcnow().strftime("%Y-%m-%d %H:%M UTC"),
        "kpis": {
            "total_students": total,
            "active_students": stats.get("active_students", 0),
            "undergraduate": stats.get("students_by_level", {}).get("undergraduate", 0),
            "postgraduate": stats.get("students_by_level", {}).get("postgraduate", 0),
            "at_risk": at_risk_count,
            "on_track_pct": round((compliance.get("green", 0) / compliance_total) * 100, 1),
            "collection_rate_pct": collection_rate,
            "total_enrollments": stats.get("total_enrollments", 0),
            "avg_gpa": leadership["gpa"]["average"],
            "international_pct": intl["international_pct"],
            "female_pct": intl["female_pct"],
            "probation_count": leadership["student_success"]["probation_or_suspended"],
            "ranking_readiness_pct": rankings_snapshot["avg_ranking_readiness_pct"] if rankings_snapshot else None,
        },
        "leadership": leadership,
        "rankings_snapshot": rankings_snapshot,
        "benchmarks": benchmarks,
        "majors": {
            "ranked": majors_sorted,
            "top": majors_sorted[:5],
            "bottom": list(reversed(majors_sorted[-5:])) if len(majors_sorted) >= 5 else list(reversed(majors_sorted)),
        },
        "nationalities_top": nationalities_top,
        "nationalities_ranked": nationalities_ranked,
        "compliance": compliance,
        "risk_by_category": risk_by_category,
        "financial_summary": {
            "total_fees_due": fees_due,
            "total_fees_paid": fees_paid,
            "total_balance_due": float(financial.get("total_balance_due") or 0),
            "total_scholarships_awarded": float(financial.get("total_scholarships_awarded") or 0),
            "payment_status": financial.get("payment_status", {}),
            "scholarship_application_status": financial.get("scholarship_application_status", {}),
        },
        "students_by_cohort": stats.get("students_by_cohort", {}),
        "students_by_gender": stats.get("students_by_gender", {}),
        "students_by_year": stats.get("students_by_year", {}),
        "insights": insights[:8],
    }


@router.get("/scholarships")
def get_scholarships(
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Open scholarship opportunities from PostgreSQL catalogue."""
    return _open_scholarships(db)


@router.get("/scholarships/my-applications")
async def get_my_scholarship_applications(
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Student applications from DB: draft, submitted for review, Awarded only."""
    from app.account_category import sync_account_category

    sync_account_category(current_user, db)
    if current_user.account_category != "student":
        raise HTTPException(status_code=403, detail="Students only")

    student_id = current_user.student_registration_number
    if not student_id:
        raise HTTPException(status_code=400, detail="No student registration number")

    schol_by_id = _scholarships_lookup(db)

    from app.student_scholarship_view import build_student_application_list_item

    rows = scholarship_db.list_student_applications(db, student_id, statuses_only=True)
    applications = [
        build_student_application_list_item(
            row, schol_by_id.get(str(row.scholarship_external_id))
        )
        for row in rows
    ]
    return {"applications": applications, "total": len(applications)}


@router.get("/scholarships/applications/{schol_id}/detail")
async def get_student_application_detail(
    schol_id: str,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Student view: application details, pipeline status, and committee review progress."""
    from app.account_category import sync_account_category
    from app.student_scholarship_view import get_student_application_detail as build_detail

    sync_account_category(current_user, db)
    if current_user.account_category != "student":
        raise HTTPException(status_code=403, detail="Students only")

    student_id = current_user.student_registration_number
    if not student_id:
        raise HTTPException(status_code=400, detail="No student registration number")

    schol_by_id = _scholarships_lookup(db)
    detail = build_detail(db, student_id, schol_id, schol_by_id.get(str(schol_id)))
    if not detail:
        raise HTTPException(status_code=404, detail="Application not found")
    return detail


@router.get("/scholarships/applications/{schol_id}/offer")
async def get_scholarship_offer(
    schol_id: str,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Stage 6: Student views their formal scholarship offer."""
    from app.account_category import sync_account_category
    from app.awards import AWARD_OFFER_SENT, build_offer_letter, expire_stale_offers

    expire_stale_offers(db)
    from app.models import ScholarshipProgram

    sync_account_category(current_user, db)
    if current_user.account_category != "student":
        raise HTTPException(status_code=403, detail="Students only")

    student_id = current_user.student_registration_number
    if not student_id:
        raise HTTPException(status_code=400, detail="No student registration number")

    row = scholarship_db.get_application(db, student_id, schol_id)
    if not row:
        raise HTTPException(status_code=404, detail="Application not found")
    if row.award_stage not in (AWARD_OFFER_SENT, "offer_accepted", "credited"):
        raise HTTPException(status_code=404, detail="No active offer for this application")

    program = (
        db.query(ScholarshipProgram)
        .filter(ScholarshipProgram.external_id == str(schol_id))
        .first()
    )
    schol_by_id = _scholarships_lookup(db)
    app_dict = scholarship_db.app_to_dict(row, schol_by_id.get(str(schol_id)))
    letter = row.offer_data or build_offer_letter(row, program)
    return {
        "application": app_dict,
        "offer_letter": letter,
        "offer_deadline": row.offer_deadline.isoformat() if row.offer_deadline else None,
        "can_respond": row.award_stage == AWARD_OFFER_SENT,
    }


@router.post("/scholarships/applications/{schol_id}/offer/accept")
async def accept_scholarship_offer(
    schol_id: str,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Stage 6–7: Student accepts offer; funds credit to tuition ledger."""
    from app.account_category import sync_account_category
    from app.awards import accept_offer, expire_stale_offers

    sync_account_category(current_user, db)
    if current_user.account_category != "student":
        raise HTTPException(status_code=403, detail="Students only")

    student_id = current_user.student_registration_number
    if not student_id:
        raise HTTPException(status_code=400, detail="No student registration number")

    expire_stale_offers(db)
    row = scholarship_db.get_application(db, student_id, schol_id)
    if not row:
        raise HTTPException(status_code=404, detail="Application not found")
    try:
        result = accept_offer(db, row)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    schol_by_id = _scholarships_lookup(db)
    return {
        **result,
        "application": scholarship_db.app_to_dict(row, schol_by_id.get(str(schol_id))),
    }


@router.post("/scholarships/applications/{schol_id}/offer/decline")
async def decline_scholarship_offer(
    schol_id: str,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Stage 6: Student declines offer; next runner-up may be notified."""
    from app.account_category import sync_account_category
    from app.awards import decline_offer
    from app.models import ScholarshipProgram

    sync_account_category(current_user, db)
    if current_user.account_category != "student":
        raise HTTPException(status_code=403, detail="Students only")

    student_id = current_user.student_registration_number
    if not student_id:
        raise HTTPException(status_code=400, detail="No student registration number")

    row = scholarship_db.get_application(db, student_id, schol_id)
    if not row:
        raise HTTPException(status_code=404, detail="Application not found")
    programs = {str(p.external_id): p for p in db.query(ScholarshipProgram).all()}
    try:
        result = decline_offer(db, row, programs=programs)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    schol_by_id = _scholarships_lookup(db)
    return {
        **result,
        "application": scholarship_db.app_to_dict(row, schol_by_id.get(str(schol_id))),
    }


@router.get("/scholarships/applications/workspace")
async def get_scholarship_workspace(
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Unified applicant dashboard data: drafts, submitted, alerts, eligible."""
    from app.account_category import sync_account_category

    sync_account_category(current_user, db)
    if current_user.account_category != "student":
        raise HTTPException(status_code=403, detail="Students only")

    student_id = current_user.student_registration_number
    if not student_id:
        raise HTTPException(status_code=400, detail="No student registration number")

    scholarships = _load_scholarship_catalog(db)
    schol_by_id = _scholarships_lookup(db)

    data = await get_my_scholarship_applications(current_user=current_user, db=db)
    apps = data.get("applications") or []

    submitted = []
    in_progress = []
    for app in apps:
        st = str(app.get("status") or "").lower()
        if st == "draft":
            in_progress.append(app)
        elif st in ("submitted for review", "awarded"):
            if st == "submitted for review":
                app["workflow_status"] = app.get("workflow_status") or "Under Triage"
            submitted.append(app)

    draft_legacy = [
        scholarship_db.draft_to_legacy_dict(r)
        for r in scholarship_db.list_student_applications(db, student_id, statuses_only=False)
        if str(r.status).lower() == "draft"
    ]

    applied_ids = {str(a.get("schol_id")) for a in apps}
    eligible = [
        s for s in scholarships
        if str(s.get("status", "")).lower() == "open" and str(s.get("id")) not in applied_ids
    ]

    alerts = build_alerts(draft_legacy, schol_by_id)

    return {
        "drafts": draft_legacy,
        "in_progress": in_progress,
        "submitted": submitted,
        "eligible": eligible,
        "alerts": alerts,
    }


@router.get("/scholarships/applications/draft/{schol_id}")
async def get_application_draft(
    schol_id: str,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    from app.account_category import sync_account_category

    sync_account_category(current_user, db)
    student_id = current_user.student_registration_number
    if not student_id:
        raise HTTPException(status_code=400, detail="No student registration number")

    scholarships = _load_scholarship_catalog(db)
    scholarship = next((s for s in scholarships if str(s.get("id")) == str(schol_id)), None)
    if not scholarship:
        raise HTTPException(status_code=404, detail="Scholarship not found")

    row = scholarship_db.get_application(db, student_id, schol_id)
    if not row:
        row = scholarship_db.upsert_draft(
            db,
            student_id,
            schol_id,
            institution_id=current_user.institution_id,
            form_data={},
            references=[],
            progress_pct=0,
        )
    draft = scholarship_db.draft_to_legacy_dict(row)

    return {
        "draft": draft,
        "scholarship": scholarship,
        "required_fields": required_field_keys(scholarship),
        "schema_type": scholarship.get("type"),
    }


@router.patch("/scholarships/applications/draft/{schol_id}")
async def patch_application_draft(
    schol_id: str,
    body: dict,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Debounced auto-save for application drafts."""
    from app.account_category import sync_account_category

    sync_account_category(current_user, db)
    student_id = current_user.student_registration_number
    if not student_id:
        raise HTTPException(status_code=400, detail="No student registration number")

    scholarships = _load_scholarship_catalog(db)
    scholarship = next((s for s in scholarships if str(s.get("id")) == str(schol_id)), None)
    if not scholarship:
        raise HTTPException(status_code=404, detail="Scholarship not found")

    refs = body.get("references")
    form_data = body.get("form_data") or {}
    progress = calc_progress(scholarship, form_data, refs or [])

    row = scholarship_db.upsert_draft(
        db,
        student_id,
        schol_id,
        institution_id=current_user.institution_id,
        form_data=form_data,
        references=refs,
        ferpa_waived=body.get("ferpa_waived"),
        progress_pct=progress,
    )
    draft = scholarship_db.draft_to_legacy_dict(row)
    return {"draft": draft, "progress_pct": progress, "saved_at": draft.get("updated_at")}


@router.post("/scholarships/applications/draft/{schol_id}/documents")
async def upload_supporting_document(
    schol_id: str,
    file: UploadFile = File(...),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Store supporting document file and attach metadata to draft form_data."""
    from app.account_category import sync_account_category

    sync_account_category(current_user, db)
    student_id = current_user.student_registration_number
    if not student_id:
        raise HTTPException(status_code=400, detail="No student registration number")

    scholarships = _load_scholarship_catalog(db)
    scholarship = next((s for s in scholarships if str(s.get("id")) == str(schol_id)), None)
    if not scholarship:
        raise HTTPException(status_code=404, detail="Scholarship not found")

    content = await file.read()
    mime = file.content_type or "application/octet-stream"
    try:
        meta = schol_docs.save_document(
            student_id,
            schol_id,
            filename=file.filename or "document",
            content=content,
            mime=mime,
        )
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc

    row = scholarship_db.get_application(db, student_id, schol_id)
    form_data = dict((row.form_data if row else None) or {})
    docs = list(form_data.get("supporting_documents") or [])
    docs.append(meta)
    form_data["supporting_documents"] = docs

    row = scholarship_db.upsert_draft(
        db,
        student_id,
        schol_id,
        institution_id=current_user.institution_id,
        form_data=form_data,
        references=(row.references_data if row else None),
        ferpa_waived=row.ferpa_waived if row else None,
        progress_pct=calc_progress(scholarship, form_data, (row.references_data if row else []) or []),
    )
    return {"document": meta, "supporting_documents": docs}


@router.delete("/scholarships/applications/draft/{schol_id}/documents/{storage_key}")
async def delete_supporting_document(
    schol_id: str,
    storage_key: str,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    from app.account_category import sync_account_category

    sync_account_category(current_user, db)
    student_id = current_user.student_registration_number
    if not student_id:
        raise HTTPException(status_code=400, detail="No student registration number")

    schol_docs.delete_document(student_id, schol_id, storage_key)
    row = scholarship_db.get_application(db, student_id, schol_id)
    if row:
        form_data = dict(row.form_data or {})
        docs = [
            d
            for d in (form_data.get("supporting_documents") or [])
            if d.get("storage_key") != storage_key
        ]
        form_data["supporting_documents"] = docs
        row.form_data = form_data
        db.commit()
    return {"ok": True, "supporting_documents": (row.form_data or {}).get("supporting_documents", []) if row else []}


@router.get("/scholarships/applications/draft/{schol_id}/documents/{storage_key}")
async def preview_supporting_document_student(
    schol_id: str,
    storage_key: str,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Student preview of own uploaded document."""
    from app.account_category import sync_account_category

    sync_account_category(current_user, db)
    student_id = current_user.student_registration_number
    if not student_id:
        raise HTTPException(status_code=400, detail="No student registration number")

    path = schol_docs.resolve_document(student_id, schol_id, storage_key)
    if not path:
        raise HTTPException(status_code=404, detail="Document not found")

    row = scholarship_db.get_application(db, student_id, schol_id)
    docs = (row.form_data or {}).get("supporting_documents") or [] if row else []
    if not any(d.get("storage_key") == storage_key for d in docs):
        raise HTTPException(status_code=403, detail="Document not part of this application")

    meta = next((d for d in docs if d.get("storage_key") == storage_key), {})
    return FileResponse(
        path,
        media_type=meta.get("mime") or "application/octet-stream",
        filename=meta.get("name") or path.name,
        headers={"Content-Disposition": f'inline; filename="{meta.get("name") or path.name}"'},
    )


@router.delete(
    "/scholarships/applications/draft/{schol_id}",
    status_code=status.HTTP_204_NO_CONTENT,
)
async def delete_application_draft(
    schol_id: str,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Discard a draft application (student only)."""
    from app.account_category import sync_account_category

    sync_account_category(current_user, db)
    student_id = current_user.student_registration_number
    if not student_id:
        raise HTTPException(status_code=400, detail="No student registration number")

    try:
        deleted = scholarship_db.delete_draft_application(db, student_id, schol_id)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc

    if not deleted:
        raise HTTPException(status_code=404, detail="Draft application not found")


@router.post("/scholarships/applications/{schol_id}/references")
async def send_reference_request(
    schol_id: str,
    body: dict,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Generate secure recommender link (mock email: returns URL in response)."""
    from app.account_category import sync_account_category

    sync_account_category(current_user, db)
    student_id = current_user.student_registration_number
    if not student_id:
        raise HTTPException(status_code=400, detail="No student registration number")

    row = scholarship_db.get_application(db, student_id, schol_id) or scholarship_db.upsert_draft(
        db, student_id, schol_id, institution_id=current_user.institution_id
    )
    draft = scholarship_db.draft_to_legacy_dict(row)
    if draft.get("ferpa_waived") is None and body.get("ferpa_waived") is None:
        raise HTTPException(status_code=400, detail="FERPA waiver choice required before sending references")

    ferpa = body.get("ferpa_waived") if body.get("ferpa_waived") is not None else draft.get("ferpa_waived")
    ref_index = int(body.get("ref_index", 0))
    recommender = {
        "name": body.get("name"),
        "title": body.get("title"),
        "institution": body.get("institution"),
        "email": body.get("email"),
        "status": "pending",
    }
    if not all([recommender["name"], recommender["email"]]):
        raise HTTPException(status_code=400, detail="Recommender name and email required")

    token = create_reference_token(student_id, schol_id, ref_index, recommender, bool(ferpa))
    refs = list(draft.get("references") or [])
    while len(refs) <= ref_index:
        refs.append({})
    refs[ref_index] = {**recommender, "token": token, "status": "pending", "sent_at": datetime.utcnow().isoformat() + "Z"}

    profile_name = current_user.full_name
    scholarship_db.upsert_draft(
        db,
        student_id,
        schol_id,
        institution_id=current_user.institution_id,
        form_data=draft.get("form_data") or {},
        references=refs,
        ferpa_waived=ferpa,
    )

    verify_path = f"/portal/recommendation/verify?token={token}"
    return {
        "message": "Reference invitation queued (mock — link returned for testing)",
        "token": token,
        "verify_path": verify_path,
        "student_name": profile_name,
        "recommender_email": recommender["email"],
    }


@router.get("/scholarships/recommendation")
async def get_recommendation_portal(token: str = Query(...)):
    """Public recommender portal — tokenized, no login."""
    rec = get_reference_by_token(token)
    if not rec:
        raise HTTPException(status_code=404, detail="Invalid or expired recommendation link")

    return {
        "student_id": rec.get("student_id"),
        "ferpa_waived": rec.get("ferpa_waived"),
        "ferpa_label": "WAIVED" if rec.get("ferpa_waived") else "NOT WAIVED",
        "recommender": rec.get("recommender"),
        "status": rec.get("status"),
        "rubric_options": ["Exceptional", "Strong", "Average", "Below Average"],
    }


@router.post("/scholarships/recommendation")
async def submit_recommendation(body: dict, db: Session = Depends(get_db)):
    """Recommender submits rating + document metadata (integrity scan simulated)."""
    token = body.get("token")
    if not token:
        raise HTTPException(status_code=400, detail="token required")

    rec = get_reference_by_token(token)
    if not rec:
        raise HTTPException(status_code=404, detail="Invalid token")
    if rec.get("status") == "completed":
        raise HTTPException(status_code=400, detail="Recommendation already submitted")

    upload = body.get("upload_meta") or {}
    allowed = {"application/pdf", "image/png", "image/jpeg"}
    mime = upload.get("mime") or "application/pdf"
    size_mb = float(upload.get("size_mb") or 0)
    if mime not in allowed:
        raise HTTPException(status_code=400, detail="Only PDF, PNG, or JPEG allowed")
    if size_mb > 10:
        raise HTTPException(status_code=400, detail="Document must be ≤ 10 MB")

    rating = body.get("rating")
    if not rating:
        raise HTTPException(status_code=400, detail="Rating required")

    upload["integrity_ok"] = True
    upload["scanned_at"] = datetime.utcnow().isoformat() + "Z"

    complete_reference(token, rating, upload)
    rec = get_reference_by_token(token)
    if rec:
        row = scholarship_db.get_application(db, rec["student_id"], rec["schol_id"])
        if row:
            refs = list(row.references_data or [])
            idx = rec["ref_index"]
            while len(refs) <= idx:
                refs.append({})
            refs[idx] = {**(rec.get("recommender") or {}), "status": "completed", "token": token}
            row.references_data = refs
            db.commit()
    return {"message": "Recommendation submitted successfully", "status": "completed"}


@router.post("/scholarships/applications/{schol_id}/submit")
async def submit_scholarship_application(
    schol_id: str,
    body: dict,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Submission gate — validates required fields, references, eligibility."""
    from app.account_category import sync_account_category

    sync_account_category(current_user, db)
    if current_user.account_category != "student":
        raise HTTPException(status_code=403, detail="Students only")

    student_id = current_user.student_registration_number
    if not student_id:
        raise HTTPException(status_code=400, detail="No student registration number")

    scholarships = _load_scholarship_catalog(db)
    scholarship = next((s for s in scholarships if str(s.get("id")) == str(schol_id)), None)
    if not scholarship:
        raise HTTPException(status_code=404, detail="Scholarship not found")

    row = scholarship_db.get_application(db, student_id, schol_id)
    form_data = body.get("form_data") or (row.form_data if row else {}) or {}
    references = (
        body.get("references")
        if body.get("references") is not None
        else (row.references_data if row else []) or []
    )
    ferpa_waived = body.get("ferpa_waived") if "ferpa_waived" in body else (row.ferpa_waived if row else None)

    require_refs = body.get("require_references", True)
    errors = validate_submission(
        scholarship, form_data, references, ferpa_waived, require_references=require_refs
    )
    if errors:
        raise HTTPException(status_code=422, detail={"errors": errors, "message": "Validation failed"})

    if row and str(row.status).lower() != "draft":
        raise HTTPException(status_code=400, detail="Application already submitted")

    try:
        gpa_val = float(form_data.get("gpa")) if form_data.get("gpa") else None
    except (TypeError, ValueError):
        gpa_val = None

    try:
        submitted_row = scholarship_db.submit_application(
            db,
            student_id,
            schol_id,
            form_data=form_data,
            references=references,
            ferpa_waived=ferpa_waived,
            gpa=gpa_val,
            institution_id=current_user.institution_id,
        )
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc

    application = scholarship_db.app_to_dict(submitted_row, scholarship)
    return {"message": "Application submitted successfully", "application": application}


@router.post("/scholarships/apply")
async def apply_for_scholarship(
    body: dict,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Submit a scholarship application for the logged-in student."""
    from app.account_category import sync_account_category

    sync_account_category(current_user, db)

    if current_user.account_category != "student":
        raise HTTPException(status_code=403, detail="Only students can apply for scholarships")

    schol_id = body.get("schol_id")
    if not schol_id:
        raise HTTPException(status_code=400, detail="schol_id is required")

    student_id = current_user.student_registration_number
    if not student_id:
        raise HTTPException(status_code=400, detail="No student registration number on account")

    scholarships = _load_scholarship_catalog(db)
    scholarship = next((s for s in scholarships if str(s.get("id")) == str(schol_id)), None)
    if not scholarship:
        raise HTTPException(status_code=404, detail="Scholarship not found")

    if str(scholarship.get("status", "")).lower() != "open":
        raise HTTPException(status_code=400, detail="Scholarship is not open for applications")

    existing = scholarship_db.get_application(db, student_id, schol_id)
    if existing and str(existing.status).lower() != "draft":
        raise HTTPException(status_code=400, detail="You have already applied for this scholarship")

    submitted_row = scholarship_db.submit_application(
        db,
        student_id,
        schol_id,
        form_data={},
        references=[],
        ferpa_waived=None,
        institution_id=current_user.institution_id,
    )
    application = scholarship_db.app_to_dict(submitted_row, scholarship)
    return {"message": "Application submitted successfully", "application": application}
