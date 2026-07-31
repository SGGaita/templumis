"""Add PG Research Tracker, PG Academic Support, and Library Resources sheets."""

from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill

ROOTS = [
    Path(__file__).resolve().parent.parent / "data",
    Path(__file__).resolve().parent / "data",
]

PG_RESEARCH_HEADERS = [
    "Student ID", "Full Name", "Programme Level", "Supervisor", "Research Area",
    "Dissertation Title", "Proposal Status", "Ethics Approval", "Data Collection",
    "Thesis Draft", "Submission Target", "Publications", "Conferences",
    "Grant Funding", "Current Challenge", "Next Milestone", "Risk Flag",
]

PG_RESEARCH_ROWS = [
    [
        "TU-PG-2022-0026", "Amara Diallo", "MSc (Coursework)", "Dr. Fatima Okello",
        "Public Health Policy", "Community Health Interventions in Rural Kenya",
        "Approved", "Cleared", "Complete", "Final draft", "2026-04-15",
        1, 2, "Internal research grant — KES 120,000", "Thesis formatting review",
        "Submit final dissertation", "Low",
    ],
    [
        "TU-PG-2023-0027", "Benjamin Kamto", "MBA (Coursework)", "Prof. James Wanjala",
        "Strategic Management", "Digital Transformation in SMEs",
        "Submitted — under review", "Pending", "In progress", "Chapter 3 draft",
        "2026-08-30", 0, 1, "None", "Supervisor feedback turnaround",
        "Complete coursework modules", "Medium",
    ],
    [
        "TU-PG-2022-0028", "Celestine Awino", "MA (Coursework)", "Dr. Grace Muthoni",
        "Education Policy", "Inclusive Education Practices in Secondary Schools",
        "Approved", "Cleared", "Complete", "Under revision", "2026-05-20",
        0, 1, "Faculty seed grant — KES 80,000", "Literature gap refinement",
        "Revise thesis chapters 2–3", "Low",
    ],
    [
        "TU-PG-2023-0029", "David Ochieng", "MSc (Research)", "Dr. Peter Kimani",
        "Data Science", "Machine Learning for Agricultural Yield Prediction",
        "Approved", "Cleared", "Active", "Chapter 4 in progress", "2026-12-01",
        1, 0, "Research council grant — KES 250,000", "Dataset licensing (FAIR compliance)",
        "Complete data analysis chapter", "Medium",
    ],
    [
        "TU-PG-2021-0030", "Elena Vasquez", "PhD", "Prof. Samuel Njoroge",
        "Climate Science", "Climate Resilience Modelling for East African Coasts",
        "Approved", "Cleared", "Active", "Draft v3", "2027-06-30",
        3, 4, "International fellowship — USD 15,000", "Publication pipeline coordination",
        "Prepare for progression review", "Low",
    ],
]

PG_SUPPORT_HEADERS = [
    "Student ID", "Service", "Provider", "Contact Email", "Status",
    "Last Session", "Next Session", "Notes", "Priority",
]

PG_SUPPORT_ROWS = [
    ["TU-PG-2022-0026", "Writing Centre", "PG Writing Lab", "writing@templumis.ac", "Active", "2026-02-10", "2026-03-05", "Thesis structure coaching", "Normal"],
    ["TU-PG-2022-0026", "Research Methods", "Graduate School", "gradschool@templumis.ac", "Completed", "2026-01-20", None, "Qualitative methods workshop", "Normal"],
    ["TU-PG-2023-0027", "Academic Coaching", "PG Support Office", "pgsupport@templumis.ac", "Scheduled", "2026-02-28", "2026-03-12", "Time management for coursework", "High"],
    ["TU-PG-2022-0028", "Library Research", "Research Librarian", "library@templumis.ac", "Active", "2026-02-15", "2026-03-08", "Systematic review search strategy", "Normal"],
    ["TU-PG-2023-0029", "Data Management", "Research Data Team", "data@templumis.ac", "Active", "2026-02-18", "2026-03-15", "FAIR data plan review", "High"],
    ["TU-PG-2021-0030", "Supervisor Meeting", "Graduate School", "gradschool@templumis.ac", "Active", "2026-02-22", "2026-03-20", "Annual progression review prep", "Normal"],
    ["TU-PG-2021-0030", "Ethics Advisory", "Research Ethics Board", "ethics@templumis.ac", "Completed", "2025-11-10", None, "Ethics clearance renewed", "Normal"],
]

LIBRARY_HEADERS = [
    "Resource ID", "Title", "Type", "Subject Area", "Programme Level",
    "FAIR — Findable (DOI/Metadata)", "FAIR — Accessible", "FAIR — Interoperable",
    "FAIR — Reusable (License)", "URL", "Description", "Access Notes",
]

LIBRARY_ROWS = [
    ["LIB-001", "TemplumIS Research Repository", "Repository", "Multidisciplinary", "All",
     "Indexed with DOI; rich metadata (Dublin Core)", "Campus SSO + ORCID login",
     "OAI-PMH API; JSON metadata export", "CC BY 4.0 for published theses", "https://repository.templumis.ac",
     "Institutional open-access repository for PG theses and datasets.", "Upload via Graduate School portal"],
    ["LIB-002", "Scopus", "Database", "Multidisciplinary", "Postgraduate",
     "Persistent IDs for indexed publications", "Campus IP + remote VPN",
     "RIS/BibTeX export; API for bibliometrics", "Publisher terms apply", "https://www.scopus.com",
     "Citation and abstract database for literature reviews.", "Use library proxy off-campus"],
    ["LIB-003", "Web of Science", "Database", "Sciences", "Postgraduate",
     "DOI-linked records", "Campus authentication", "EndNote integration", "Publisher license", "https://www.webofscience.com",
     "Peer-reviewed journal index and citation tracking.", "Book a librarian session for advanced search"],
    ["LIB-004", "Kenya National Research Data Archive", "Dataset", "Social Sciences", "Postgraduate",
     "Registered dataset DOIs; README metadata", "Open download + API key for restricted sets",
     "CSV, JSON, STATA interoperable formats", "CC BY 4.0 / custom data agreements", "https://data.kenya-research.ac",
     "National FAIR-compliant datasets for policy and health research.", "Register before downloading restricted data"],
    ["LIB-005", "IEEE Xplore", "Database", "Engineering & CS", "Postgraduate",
     "IEEE DOI metadata", "Institutional subscription", "BibTeX, XML exports", "IEEE terms of use", "https://ieeexplore.ieee.org",
     "Engineering and computing literature for MSc/PhD research.", "Link ORCID for author profiles"],
    ["LIB-006", "JSTOR", "Database", "Humanities & Social Sciences", "All",
     "Stable URLs and citation metadata", "Campus login", "PDF + citation export", "JSTOR license", "https://www.jstor.org",
     "Archive of journals and books for humanities PG research.", "Use advanced search filters"],
    ["LIB-007", "Zotero Research Guide", "Tool", "Research Skills", "Postgraduate",
     "Guide metadata on library catalogue", "Open access guide", "RIS/BibTeX compatible", "Creative Commons guide", "https://libguides.templumis.ac/zotero",
     "Reference management and FAIR citation workflows.", "Install desktop + browser connector"],
    ["LIB-008", "Open Data Kenya — Agriculture", "Dataset", "Agriculture", "Postgraduate",
     "Dataset catalogue with DOIs", "Open API + bulk download", "CSV, GeoJSON", "Open Government Licence", "https://opendata.templumis.ac/agriculture",
     "FAIR agricultural datasets for MSc research projects.", "Cite dataset DOI in thesis"],
]


def _style_header(ws, headers):
    fill = PatternFill("solid", fgColor="1E3A5F")
    font = Font(bold=True, color="FFFFFF", size=10)
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = fill
        cell.font = font


def seed_workbook(path: Path):
    wb = openpyxl.load_workbook(path)
    for sheet_name in ("PG Research Tracker", "PG Academic Support", "Library Resources"):
        if sheet_name in wb.sheetnames:
            del wb[sheet_name]

    ws_r = wb.create_sheet("PG Research Tracker")
    _style_header(ws_r, PG_RESEARCH_HEADERS)
    for row in PG_RESEARCH_ROWS:
        ws_r.append(row)

    ws_s = wb.create_sheet("PG Academic Support")
    _style_header(ws_s, PG_SUPPORT_HEADERS)
    for row in PG_SUPPORT_ROWS:
        ws_s.append(row)

    ws_l = wb.create_sheet("Library Resources")
    _style_header(ws_l, LIBRARY_HEADERS)
    for row in LIBRARY_ROWS:
        ws_l.append(row)

    wb.save(path)
    print(f"Seeded {path.name} ({path})")


def main():
    for root in ROOTS:
        for name in ("templumis_university.xlsx", "templumis_university_v2.xlsx"):
            path = root / name
            if path.exists():
                seed_workbook(path)


if __name__ == "__main__":
    main()
