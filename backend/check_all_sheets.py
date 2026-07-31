import openpyxl

# Load the Excel file
wb = openpyxl.load_workbook('data/templumis_university.xlsx', data_only=True)

print("All sheets in the workbook:")
print("="*60)

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f"\n📊 Sheet: {sheet_name}")
    print(f"   Dimensions: {ws.dimensions}")
    print(f"   First row: {[cell.value for cell in ws[1]]}")
    
    # Check if it might be rankings related
    first_cell = ws['A1'].value
    if first_cell and ('ranking' in str(first_cell).lower() or 'indicator' in str(first_cell).lower()):
        print(f"   ⭐ Potential rankings sheet detected!")
        print(f"   First 10 rows:")
        for idx, row in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=True), 1):
            print(f"      Row {idx}: {row}")
